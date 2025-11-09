from flask import Flask, render_template, request, redirect, url_for, session, flash, send_file, jsonify
import datetime
import os
import io
from werkzeug.utils import secure_filename
import mysql.connector
import openpyxl
from openpyxl.styles import Font, Alignment
from utils import get_department_stats, get_gender_stats, get_appointment_stats, get_experience_stats, get_designation_stats

# Add these constants and functions at the top
ALLOWED_EXTENSIONS = {'pdf', 'doc', 'docx', 'jpg', 'jpeg', 'png'}
MAX_FILE_SIZE = 5 * 1024 * 1024  # 5MB

def get_user_role():
    """Get current user's role with new role names"""
    role = session.get('role', 'Faculty')
    print(f"🔍 DEBUG get_user_role(): session role = '{role}'")
    # Map any old roles to new names for backward compatibility
    role_mapping = {
        'admin': 'IQAC',
        'editor': 'Office', 
        'viewer': 'Faculty',
        'Admin': 'IQAC',
        'Editor': 'Office',
        'Viewer': 'Faculty',
        'IQAC(admin)': 'IQAC'
    }
    final_role = role_mapping.get(role, role)
    print(f"🔍 DEBUG get_user_role(): final role = '{final_role}'")
    return final_role

def can_edit_faculty():
    """Check if user can edit faculty data"""
    return get_user_role() in ['IQAC', 'Office']

def can_delete_faculty():
    """Check if user can delete faculty"""
    return get_user_role() == 'IQAC'

def can_add_faculty():
    """Check if user can add faculty"""
    return get_user_role() in ['IQAC', 'Office']

def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

app = Flask(__name__)
app.secret_key = 'faculty-secret-key'

def get_db_connection():
    try:
        conn = mysql.connector.connect(
            host=os.environ['MYSQLHOST'],
            user=os.environ['MYSQLUSER'],
            password=os.environ['MYSQLPASSWORD'],
            database=os.environ['MYSQLDATABASE'],
            port=int(os.environ['MYSQLPORT']),
            connect_timeout=30,
            autocommit=True
        )
        print("✅ Database connected successfully!")
        return conn
    except mysql.connector.Error as e:
        print(f"❌ Database connection failed: {e}")
        print(f"   Host: {os.environ.get('MYSQLHOST')}")
        print(f"   Database: {os.environ.get('MYSQLDATABASE')}")
        return None
    except KeyError as e:
        print(f"❌ Missing environment variable: {e}")
        return None
def login_required(f):
    """Decorator to require login for routes"""
    from functools import wraps
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get('logged_in'):
            return redirect('/login')
        return f(*args, **kwargs)
    return decorated_function    

def can_edit_publications(faculty_id):
    """Check if current user can edit publications for this faculty"""
    user_role = get_user_role()
    user_email = session.get('email', '')
    
    # IQAC and Office can view but not edit others' publications
    if user_role in ['IQAC', 'Office']:
        # Check if this is their own faculty record
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute('SELECT email FROM faculty WHERE id = %s', (faculty_id,))
        faculty = cursor.fetchone()
        cursor.close()
        conn.close()
        
        if faculty and faculty['email'] == user_email:
            return True  # Can edit their own
        return False    # Cannot edit others'
    
    # Faculty can only edit their own
    elif user_role == 'Faculty':
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute('SELECT email FROM faculty WHERE id = %s', (faculty_id,))
        faculty = cursor.fetchone()
        cursor.close()
        conn.close()
        
        return faculty and faculty['email'] == user_email
    
    return False

def check_publication_access(faculty_id):
    """Check if current user can edit publications for this faculty"""
    user_role = get_user_role()
    user_email = session.get('email', '')
    
    # IQAC(admin) and Office can view but not edit others' publications
    if user_role in ['IQAC(admin)', 'Office']:
        # Check if this is their own faculty record
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute('SELECT email FROM faculty WHERE id = %s', (faculty_id,))
        faculty = cursor.fetchone()
        cursor.close()
        conn.close()
        
        if faculty and faculty['email'] == user_email:
            return True  # Can edit their own
        return False    # Cannot edit others'
    
    # Faculty can only edit their own
    elif user_role == 'Faculty':
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute('SELECT email FROM faculty WHERE id = %s', (faculty_id,))
        faculty = cursor.fetchone()
        cursor.close()
        conn.close()
        
        return faculty and faculty['email'] == user_email
    
    return False      

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username'].strip()
        email = request.form['email'].strip()
        password = request.form['password']
        
        print(f"🔍 LOGIN ATTEMPT: username='{username}', email='{email}'")
        
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        try:
            # ✅ REQUIRE BOTH USERNAME AND EMAIL TO MATCH
            cursor.execute('SELECT * FROM users WHERE username = %s AND email = %s AND password_hash = %s', 
                          (username, email, password))
            user = cursor.fetchone()
            
            if user:
                if not user['approved']:
                    print(f"🔍 LOGIN FAILED: User '{user['username']}' not approved")
                    cursor.close()
                    conn.close()
                    flash('⏳ Account pending admin approval. Please wait for IQAC approval.', 'error')
                    return render_template('login.html', error='⏳ Account pending admin approval. Please wait for IQAC approval.')
                
                # ✅ SUCCESSFUL LOGIN
                session['user_id'] = user['id']
                session['username'] = user['username']
                session['email'] = user['email']
                session['role'] = user['role']
                session['logged_in'] = True
                
                # Update last login
                cursor.execute('UPDATE users SET last_login = NOW() WHERE id = %s', (user['id'],))
                conn.commit()
                
                print(f"🔍 LOGIN SUCCESS: User '{user['username']}' logged in as '{user['role']}'")
                cursor.close()
                conn.close()
                
                flash(f'✅ Welcome back, {user["username"]}!', 'success')
                return redirect('/')
            else:
                # ✅ CHECK WHAT WENT WRONG FOR BETTER ERROR MESSAGES
                cursor.execute('SELECT username, email FROM users WHERE username = %s AND email = %s', 
                              (username, email))
                user_exists = cursor.fetchone()
                
                if user_exists:
                    # Username and email match but wrong password
                    print(f"🔍 LOGIN FAILED: Wrong password for user '{username}'")
                    cursor.close()
                    conn.close()
                    flash('❌ Invalid password. Please try again.', 'error')
                    return render_template('login.html', error='❌ Invalid password. Please try again.', 
                                         form_data={'username': username, 'email': email})
                else:
                    # Check if username exists but email doesn't match
                    cursor.execute('SELECT username FROM users WHERE username = %s', (username,))
                    username_exists = cursor.fetchone()
                    
                    cursor.execute('SELECT email FROM users WHERE email = %s', (email,))
                    email_exists = cursor.fetchone()
                    
                    cursor.close()
                    conn.close()
                    
                    if username_exists and email_exists:
                        error_msg = '❌ Username and email combination is incorrect.'
                    elif username_exists:
                        error_msg = '❌ Email does not match this username.'
                    elif email_exists:
                        error_msg = '❌ Username does not match this email.'
                    else:
                        error_msg = '❌ Username and email not found.'
                    
                    print(f"🔍 LOGIN FAILED: {error_msg}")
                    flash(error_msg, 'error')
                    return render_template('login.html', error=error_msg, 
                                         form_data={'username': username, 'email': email})
                    
        except Exception as e:
            print(f"🔍 LOGIN ERROR: {str(e)}")
            if 'conn' in locals() and conn.is_connected():
                cursor.close()
                conn.close()
            flash('❌ System error. Please try again later.', 'error')
            return render_template('login.html', error='❌ System error. Please try again later.',
                                 form_data={'username': username, 'email': email})
    
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect('/login')

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        username = request.form['username'].strip()
        email = request.form['email'].lower().strip()
        password = request.form['password']
        role = request.form.get('role', 'Faculty')
        
        print(f"🔍 REGISTRATION DEBUG: Starting registration for {username} ({email})")
        
        # Validate inputs
        if not username or not email or not password:
            flash('❌ All fields are required!', 'error')
            return render_template('register.html', form_data=request.form)
        
        approved = False
        
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        try:
            # ✅ ONLY CHECK: Duplicate EMAIL (email must be unique)
            cursor.execute('SELECT id, username FROM users WHERE email = %s', (email,))
            existing_email = cursor.fetchone()
            
            if existing_email:
                print(f"🔍 REGISTRATION DEBUG: Duplicate email found for {email}")
                cursor.close()
                conn.close()
                flash(f'❌ Email "{email}" is already registered. Please use a different email address.', 'error')
                return render_template('register.html', form_data=request.form)
            
            # ✅ INSERT NEW USER (allow same username with different email)
            cursor.execute(
                'INSERT INTO users (username, email, password_hash, role, approved, created_at) VALUES (%s, %s, %s, %s, %s, NOW())',
                (username, email, password, role, approved)
            )
            conn.commit()
            
            user_id = cursor.lastrowid
            print(f"🔍 REGISTRATION DEBUG: Successfully registered user ID {user_id}")
            
            cursor.close()
            conn.close()
            
            # ✅ SUCCESS - Show message on SAME PAGE instead of redirecting immediately
            flash('✅ Registration submitted successfully! Please wait for admin approval.', 'success')
            return render_template('register.html', form_data={}, show_success=True)
                
        except mysql.connector.Error as err:
            print(f"🔍 REGISTRATION DEBUG: MySQL Error - {err}")
            
            if 'conn' in locals() and conn.is_connected():
                conn.rollback()
                cursor.close()
                conn.close()
            
            # Handle specific MySQL errors
            if err.errno == 1062:  # Duplicate entry
                flash('❌ This email is already registered. Please use a different email.', 'error')
            else:
                flash(f'❌ Database error: {str(err)}', 'error')
            
            return render_template('register.html', form_data=request.form)
            
        except Exception as e:
            print(f"🔍 REGISTRATION DEBUG: General Error - {e}")
            
            if 'conn' in locals() and conn.is_connected():
                conn.rollback()
                cursor.close()
                conn.close()
                
            flash(f'❌ Unexpected error: {str(e)}', 'error')
            return render_template('register.html', form_data=request.form)
    
    print("🔍 REGISTRATION DEBUG: GET request for registration form")
    return render_template('register.html', form_data={})

@app.route('/')
@login_required
def index():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    
    # BASIC STATISTICS - FOR ALL USERS
    cursor.execute('SELECT COUNT(*) as total FROM faculty')
    total_faculty = cursor.fetchone()['total']
    
    cursor.execute("SELECT COUNT(*) as regular FROM faculty WHERE appointment_type = 'Regular'")
    regular_faculty = cursor.fetchone()['regular']
    
    cursor.execute('SELECT COUNT(DISTINCT department) as depts FROM faculty')
    total_departments = cursor.fetchone()['depts']
    
    # Get faculty data for statistics
    if get_user_role() in ['Faculty']:
        # Faculty can only see their own data
        cursor.execute('SELECT * FROM faculty WHERE email = %s', (session.get('email'),))
    else:
        # IQAC/Office can see all faculty
        cursor.execute('SELECT * FROM faculty')
    
    faculty_data = cursor.fetchall()
    
    # Get detailed designation counts
    cursor.execute("SELECT COUNT(*) as count FROM faculty WHERE designation = 'Professor'")
    professor_count = cursor.fetchone()['count']
    
    cursor.execute("SELECT COUNT(*) as count FROM faculty WHERE designation = 'Associate Professor'")
    associate_professor_count = cursor.fetchone()['count']
    
    cursor.execute("SELECT COUNT(*) as count FROM faculty WHERE designation = 'Assistant Professor'")
    assistant_professor_count = cursor.fetchone()['count']
    
    # Get qualification counts
    cursor.execute("SELECT COUNT(DISTINCT q.faculty_id) as phd_count FROM qualifications q WHERE q.qualification_type = 'Ph.D' AND q.highest_degree = 1")
    phd_count = cursor.fetchone()['phd_count']
    
    cursor.execute("SELECT COUNT(DISTINCT q.faculty_id) as pg_count FROM qualifications q WHERE q.qualification_type IN ('PG', 'Post Graduate', 'M.Tech', 'M.E', 'M.Sc', 'M.A', 'M.Com') AND q.highest_degree = 1")
    pg_count = cursor.fetchone()['pg_count']
    
    # Generate statistics with FIXED logic
    designation_stats = get_designation_stats(faculty_data)
    gender_stats = get_gender_stats(faculty_data)
    appointment_stats = get_appointment_stats(faculty_data)

    # FIXED: Calculate experience stats with consistent logic
    experience_stats = []
    if faculty_data:
        count_0_5 = sum(1 for f in faculty_data if f.get('overall_exp', 0) <= 5.9)
        count_6_10 = sum(1 for f in faculty_data if 6 <= f.get('overall_exp', 0) <= 10.9)
        count_10_plus = sum(1 for f in faculty_data if f.get('overall_exp', 0) > 10.9)
    
        experience_stats = [
        {'experience_category': '0-5', 'count': count_0_5},
        {'experience_category': '6-10', 'count': count_6_10},
        {'experience_category': '10+', 'count': count_10_plus}
    ]
    else:
        experience_stats = get_experience_stats(faculty_data)
    
    # R&D Publications Statistics - ONLY for IQAC and Office
    journal_count = 0
    conference_count = 0
    book_chapter_count = 0
    patent_count = 0
    book_count = 0
    total_publications = 0
    
    # Only calculate R&D stats for IQAC and Office roles
    user_role = get_user_role()
    
    if user_role in ['IQAC', 'Office']:
        try:
            cursor.execute('SELECT COUNT(*) as count FROM journal_publications')
            journal_count = cursor.fetchone()['count']
            
            cursor.execute('SELECT COUNT(*) as count FROM conference_publications')
            conference_count = cursor.fetchone()['count']
            
            cursor.execute('SELECT COUNT(*) as count FROM book_chapters')
            book_chapter_count = cursor.fetchone()['count']
            
            cursor.execute('SELECT COUNT(*) as count FROM patents')
            patent_count = cursor.fetchone()['count']
            
            cursor.execute('SELECT COUNT(DISTINCT book_title) as count FROM book_chapters')
            book_count = cursor.fetchone()['count']
            
            total_publications = journal_count + conference_count + book_chapter_count + patent_count
            
        except Exception as e:
            print(f"R&D Statistics Error: {e}")
    
    cursor.close()
    conn.close()
    
    return render_template('index.html',
                         total_faculty=total_faculty,
                         regular_faculty=regular_faculty,
                         total_departments=total_departments,
                         professor_count=professor_count,
                         associate_professor_count=associate_professor_count,
                         assistant_professor_count=assistant_professor_count,
                         phd_count=phd_count,
                         pg_count=pg_count,
                         designation_stats=designation_stats,
                         gender_stats=gender_stats,
                         appointment_stats=appointment_stats,
                         experience_stats=experience_stats,
                         journal_count=journal_count,
                         conference_count=conference_count,
                         book_chapter_count=book_chapter_count,
                         patent_count=patent_count,
                         book_count=book_count,
                         total_publications=total_publications)

@app.route('/faculty')
@login_required
def faculty_list():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    
    # Get search and filter parameters
    search = request.args.get('search', '')
    department = request.args.get('department', '')
    appointment_type = request.args.get('appointment_type', '')
    exp_from = request.args.get('exp_from', '')
    exp_to = request.args.get('exp_to', '')
    designation = request.args.get('designation', '')
    
    print(f"🔍 FACULTY_LIST - RAW FILTERS:")
    print(f"   - search: '{search}'")
    print(f"   - department: '{department}'") 
    print(f"   - designation: '{designation}'")
    print(f"   - appointment_type: '{appointment_type}'")
    print(f"   - exp_from: '{exp_from}'")
    print(f"   - exp_to: '{exp_to}'")
    
    # 🔒 ROLE-BASED DATA ACCESS
    if get_user_role() in ['Faculty']:
        # Faculty can only see their own data
        query = 'SELECT * FROM faculty WHERE email = %s'
        params = [session.get('email')]
    else:
        # IQAC(admin)/Office can see all faculty
        query = 'SELECT * FROM faculty WHERE 1=1'
        params = []
    
    # Add filters for Faculty too (but only for their own data)
    if search and search.strip():
        query += ' AND (name_ssc LIKE %s OR employee_id LIKE %s)'
        params.extend([f'%{search}%', f'%{search}%'])
    
    if department and department.strip():
        query += ' AND department = %s'
        params.append(department)
    
    if designation and designation.strip():
        query += ' AND designation = %s'
        params.append(designation)
    
    if appointment_type and appointment_type.strip():
        query += ' AND appointment_type = %s'
        params.append(appointment_type)
    
    # FIXED: Experience range filter - Handle decimal values properly
    if exp_from and exp_from.strip():
        try:
            query += ' AND overall_exp >= %s'
            params.append(float(exp_from))
            print(f"🔍 DEBUG: Added exp_from filter: overall_exp >= {exp_from}")
        except ValueError:
            print(f"⚠️ Invalid exp_from value: {exp_from}")
    
    if exp_to and exp_to.strip():
        try:
            query += ' AND overall_exp <= %s'
            params.append(float(exp_to))
            print(f"🔍 DEBUG: Added exp_to filter: overall_exp <= {exp_to}")
        except ValueError:
            print(f"⚠️ Invalid exp_to value: {exp_to}")
    
    query += ' ORDER BY name_ssc'
    
    print(f"📊 FACULTY_LIST - FINAL QUERY: {query}")
    print(f"🔧 FACULTY_LIST - QUERY PARAMS: {params}")
    
    cursor.execute(query, params)
    faculty = cursor.fetchall()
    
    print(f"✅ FACULTY_LIST - FOUND {len(faculty)} RECORDS")
    if faculty:
        for f in faculty:
            print(f"   - {f['employee_id']}: {f['name_ssc']} | Dept: {f['department']} | Exp: {f.get('overall_exp', 'N/A')}")
    else:
        print("   - No records found")
    
    cursor.close()
    conn.close()
    
    return render_template('faculty_list.html', 
                         faculty=faculty, 
                         user_role=get_user_role())

@app.route('/add_faculty', methods=['GET', 'POST'])
@login_required
def add_faculty():
    if request.method == 'POST':
        try:
            # Get form data first for duplicate checks
            employee_id = request.form['employee_id']
            email = request.form['email']
            
            # ✅ CHECK FOR DUPLICATE EMPLOYEE ID BEFORE PROCESSING
            conn = get_db_connection()
            cursor = conn.cursor(dictionary=True)
            cursor.execute('SELECT id, name_ssc FROM faculty WHERE employee_id = %s', (employee_id,))
            existing_employee = cursor.fetchone()
            if existing_employee:
                cursor.close()
                conn.close()
                flash(f'❌ Employee ID "{employee_id}" already assigned to {existing_employee["name_ssc"]}. Please use a different Employee ID.', 'error')
                return render_template('add_faculty.html', form_data=request.form)
            
            # ✅ CHECK FOR DUPLICATE EMAIL
            cursor.execute('SELECT id, name_ssc FROM faculty WHERE email = %s', (email,))
            existing_email = cursor.fetchone()
            if existing_email:
                cursor.close()
                conn.close()
                flash(f'❌ Email "{email}" already registered for {existing_email["name_ssc"]}. Please use a different email address.', 'error')
                return render_template('add_faculty.html', form_data=request.form)
            
            cursor.close()
            conn.close()

            # Handle photo upload
            photo_path = None
            if 'photo' in request.files:
                photo = request.files['photo']
                if photo and photo.filename != '':
                    if allowed_file(photo.filename):
                        # Create uploads directory if not exists
                        upload_folder = 'static/uploads/photos'
                        os.makedirs(upload_folder, exist_ok=True)
                        
                        # Generate unique filename
                        filename = secure_filename(photo.filename)
                        unique_filename = f"{request.form['employee_id']}_{filename}"
                        photo_path = os.path.join(upload_folder, unique_filename)
                        photo.save(photo_path)

                        # Store relative path for web access
                        photo_path = f"uploads/photos/{unique_filename}"
                    else:
                        flash('❌ Invalid photo format. Please use JPG, PNG, or JPEG files.', 'error')
                        return render_template('add_faculty.html', form_data=request.form)

            # Handle document upload
            document_path = None
            if 'name_change_proof' in request.files:
                document = request.files['name_change_proof']
                if document and document.filename != '':
                    if allowed_file(document.filename) and document.content_length <= MAX_FILE_SIZE:
                        # Create documents upload directory
                        doc_upload_folder = 'static/uploads/documents'
                        os.makedirs(doc_upload_folder, exist_ok=True)
                        
                        # Secure filename and save
                        doc_filename = secure_filename(document.filename)
                        unique_docname = f"{request.form['employee_id']}_proof_{doc_filename}"
                        doc_save_path = os.path.join(doc_upload_folder, unique_docname)
                        document.save(doc_save_path)
                        
                        # Store relative path
                        document_path = f"uploads/documents/{unique_docname}"
                    elif document.content_length > MAX_FILE_SIZE:
                        flash('❌ File size too large. Maximum 5MB allowed.', 'error')
                        return render_template('add_faculty.html', form_data=request.form)
                    else:
                        flash('❌ Invalid document format. Please use PDF, DOC, DOCX, JPG, or PNG files.', 'error')
                        return render_template('add_faculty.html', form_data=request.form)

            # Get all form data including new fields
            employee_id = request.form['employee_id']
            name_ssc = request.form['name_ssc']
            name_change = 'name_change' in request.form
            blood_group = request.form.get('blood_group') or None
            alternative_mobile = request.form.get('alternative_mobile') or None
            bank_name = request.form['bank_name']

            # Validate and get date fields
            dob = request.form['dob']
            if not dob:
                flash('❌ Date of Birth is required', 'error')
                return render_template('add_faculty.html', form_data=request.form)
            
            try:
                datetime.datetime.strptime(dob, '%Y-%m-%d')
            except ValueError:
                flash('❌ Invalid Date of Birth format. Please use YYYY-MM-DD format', 'error')
                return render_template('add_faculty.html', form_data=request.form)
                
            gender = request.form['gender']
            marital_status = request.form.get('marital_status', '')
            father_name = request.form['father_name']
            present_address = request.form['present_address']
            permanent_address = request.form['permanent_address']
            email = request.form['email']
            mobile_no = request.form['mobile_no']
            department = request.form['department']
            designation = request.form['designation']
            
            # Validate date of joining
            date_of_joining = request.form['date_of_joining']
            if not date_of_joining:
                flash('❌ Date of Joining is required', 'error')
                return render_template('add_faculty.html', form_data=request.form)
            
            try:
                datetime.datetime.strptime(date_of_joining, '%Y-%m-%d')
            except ValueError:
                flash('❌ Invalid Date of Joining format. Please use YYYY-MM-DD format', 'error')
                return render_template('add_faculty.html', form_data=request.form)
                
            appointment_type = request.form['appointment_type']
            aadhaar_number = request.form.get('aadhaar_number', '')
            pan_number = request.form.get('pan_number', '')
            bank_account_no = request.form['bank_account_no']
            ifsc_code = request.form['ifsc_code']
            caste = request.form['caste']
            subcaste = request.form.get('subcaste', '')
            ratified = request.form.get('ratified', 'No')
            ratified_designation = request.form.get('ratified_designation', '')
            
            # Handle optional dates
            ratification_date = request.form.get('ratification_date') or None
            if ratification_date:
                try:
                    datetime.datetime.strptime(ratification_date, '%Y-%m-%d')
                except ValueError:
                    flash('❌ Invalid Ratification Date format. Please use YYYY-MM-DD format', 'error')
                    return render_template('add_faculty.html', form_data=request.form)
                    
            previous_employment_date = request.form.get('previous_employment_date') or None
            if previous_employment_date:
                try:
                    datetime.datetime.strptime(previous_employment_date, '%Y-%m-%d')
                except ValueError:
                    flash('❌ Invalid Previous Employment Date format. Please use YYYY-MM-DD format', 'error')
                    return render_template('add_faculty.html', form_data=request.form)
                    
            resignation_date = request.form.get('resignation_date') or None
            if resignation_date:
                try:
                    datetime.datetime.strptime(resignation_date, '%Y-%m-%d')
                except ValueError:
                    flash('❌ Invalid Resignation Date format. Please use YYYY-MM-DD format', 'error')
                    return render_template('add_faculty.html', form_data=request.form)
            
            # Experience fields - USE FRONTEND CALCULATIONS
            teaching_exp_pragati = float(request.form.get('teaching_exp_pragati', 0))
            teaching_exp_other = float(request.form.get('teaching_exp_other', 0))
            industrial_exp = float(request.form.get('industrial_exp', 0))
            overall_exp = float(request.form.get('overall_exp', 0))

            # FIXED: Auto-calculate experience category with correct ranges
            if overall_exp <= 5.9:
                experience_category = '0-5'
            elif overall_exp <= 10.9:
                experience_category = '6-10'
            else:
                experience_category = '10+'

            # Validate that the frontend calculations are reasonable
            calculated_total = teaching_exp_pragati + teaching_exp_other + industrial_exp
            if abs(calculated_total - overall_exp) > 0.1:
                # Allow small rounding differences
                flash('❌ Experience calculation mismatch detected. Please refresh and try again.', 'error')
                return render_template('edit_faculty.html', faculty=request.form)

            # Insert into database
            conn = get_db_connection()
            cursor = conn.cursor()
            cursor.execute(
                '''INSERT INTO faculty 
                (employee_id, name_ssc, name_change, name_change_proof, dob, gender, blood_group, marital_status, 
                 father_name, present_address, permanent_address, email, mobile_no, alternative_mobile, department, 
                 designation, date_of_joining, appointment_type, aadhaar_number, pan_number, 
                 bank_name, bank_account_no, ifsc_code, photo_path, experience_category, caste, subcaste, 
                 ratified, ratified_designation, ratification_date, previous_employment_date, resignation_date,
                 teaching_exp_pragati, teaching_exp_other, industrial_exp, overall_exp) 
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)''',
                (employee_id, name_ssc, name_change, document_path, dob, gender, blood_group, marital_status,
                 father_name, present_address, permanent_address, email, mobile_no, alternative_mobile, department, 
                 designation, date_of_joining, appointment_type, aadhaar_number, pan_number,
                 bank_name, bank_account_no, ifsc_code, photo_path, experience_category, caste, subcaste,
                 ratified, ratified_designation, ratification_date, previous_employment_date, resignation_date,
                 teaching_exp_pragati, teaching_exp_other, industrial_exp, overall_exp)
            )
            conn.commit()
            cursor.close()
            conn.close()
            
            flash('✅ Faculty member added successfully!', 'success')
            return redirect('/faculty')
            
        except mysql.connector.Error as err:
            # Handle database errors
            if 'conn' in locals() and conn.is_connected():
                conn.rollback()
                cursor.close()
                conn.close()
            
            error_messages = {
                1062: "❌ Duplicate entry detected. Employee ID or Email already exists.",
                1452: "❌ Reference error. Please check department or other related data.",
                1406: "❌ Data too long for one or more fields.",
                1366: "❌ Incorrect data format in one or more fields.",
                1048: "❌ Required field is missing. Please check all mandatory fields."
            }
            
            user_message = error_messages.get(err.errno, f'❌ Database error: {str(err)}')
            flash(user_message, 'error')
            return render_template('add_faculty.html', form_data=request.form)
            
        except Exception as e:
            # Handle any other unexpected errors
            flash(f'❌ Unexpected error occurred: {str(e)}', 'error')
            return render_template('add_faculty.html', form_data=request.form)
    
    return render_template('add_faculty.html')

@app.route('/edit_faculty/<int:faculty_id>', methods=['GET', 'POST'])
@login_required
def edit_faculty(faculty_id):
    print(f"DEBUG: Edit faculty route accessed for ID: {faculty_id}")
    
    if request.method == 'POST':
        try:
            # Get form data first for duplicate checks
            employee_id = request.form['employee_id']
            email = request.form['email']
            
            # ✅ CHECK FOR DUPLICATE EMPLOYEE ID (excluding current faculty)
            conn = get_db_connection()
            cursor = conn.cursor(dictionary=True)
            cursor.execute('SELECT id, name_ssc FROM faculty WHERE employee_id = %s AND id != %s', (employee_id, faculty_id))
            existing_employee = cursor.fetchone()
            if existing_employee:
                cursor.close()
                conn.close()
                flash(f'❌ Employee ID "{employee_id}" already assigned to {existing_employee["name_ssc"]}. Please use a different Employee ID.', 'error')
                return render_template('edit_faculty.html', faculty=request.form)
            
            # ✅ CHECK FOR DUPLICATE EMAIL (excluding current faculty)
            cursor.execute('SELECT id, name_ssc FROM faculty WHERE email = %s AND id != %s', (email, faculty_id))
            existing_email = cursor.fetchone()
            if existing_email:
                cursor.close()
                conn.close()
                flash(f'❌ Email "{email}" already registered for {existing_email["name_ssc"]}. Please use a different email address.', 'error')
                return render_template('edit_faculty.html', faculty=request.form)
            
            cursor.close()
            conn.close()

            # Get ALL form data including file uploads
            employee_id = request.form['employee_id']
            name_ssc = request.form['name_ssc']
            name_change = 'name_change' in request.form
            dob = request.form['dob']
            
            # Validate date format for dob
            try:
                datetime.datetime.strptime(dob, '%Y-%m-%d')
            except ValueError:
                flash('❌ Invalid Date of Birth format. Please use YYYY-MM-DD format', 'error')
                return render_template('edit_faculty.html', faculty=request.form)
                
            gender = request.form['gender']
            blood_group = request.form.get('blood_group') or None
            marital_status = request.form.get('marital_status', '')
            father_name = request.form['father_name']
            present_address = request.form['present_address']
            permanent_address = request.form['permanent_address']
            email = request.form['email']
            mobile_no = request.form['mobile_no']
            alternative_mobile = request.form.get('alternative_mobile') or None
            department = request.form['department']
            designation = request.form['designation']
            date_of_joining = request.form['date_of_joining']
            
            # Validate date of joining
            try:
                datetime.datetime.strptime(date_of_joining, '%Y-%m-%d')
            except ValueError:
                flash('❌ Invalid Date of Joining format. Please use YYYY-MM-DD format', 'error')
                return render_template('edit_faculty.html', faculty=request.form)
                
            appointment_type = request.form['appointment_type']
            aadhaar_number = request.form.get('aadhaar_number', '')
            pan_number = request.form.get('pan_number', '')
            bank_name = request.form['bank_name']
            bank_account_no = request.form['bank_account_no']
            ifsc_code = request.form['ifsc_code']
            caste = request.form['caste']
            subcaste = request.form.get('subcaste', '')
            ratified = request.form.get('ratified', 'No')
            ratified_designation = request.form.get('ratified_designation', '')
            
            # Handle optional dates with validation
            ratification_date = request.form.get('ratification_date') or None
            if ratification_date:
                try:
                    datetime.datetime.strptime(ratification_date, '%Y-%m-%d')
                except ValueError:
                    flash('❌ Invalid Ratification Date format. Please use YYYY-MM-DD format', 'error')
                    return render_template('edit_faculty.html', faculty=request.form)
                    
            previous_employment_date = request.form.get('previous_employment_date') or None
            if previous_employment_date:
                try:
                    datetime.datetime.strptime(previous_employment_date, '%Y-%m-%d')
                except ValueError:
                    flash('❌ Invalid Previous Employment Date format. Please use YYYY-MM-DD format', 'error')
                    return render_template('edit_faculty.html', faculty=request.form)
                    
            resignation_date = request.form.get('resignation_date') or None
            if resignation_date:
                try:
                    datetime.datetime.strptime(resignation_date, '%Y-%m-%d')
                except ValueError:
                    flash('❌ Invalid Resignation Date format. Please use YYYY-MM-DD format', 'error')
                    return render_template('edit_faculty.html', faculty=request.form)
            
            # Experience fields
            teaching_exp_pragati = float(request.form.get('teaching_exp_pragati', 0))
            teaching_exp_other = float(request.form.get('teaching_exp_other', 0))
            industrial_exp = float(request.form.get('industrial_exp', 0))
            overall_exp = float(request.form.get('overall_exp', 0))
            
            # Auto-calculate experience category
            if overall_exp <= 5.9:
                experience_category = '0-5'
            elif overall_exp <= 10.9:
                experience_category = '6-10'
            else:
                experience_category = '10+'

            # Handle photo upload with validation
            photo_path = None
            if 'photo' in request.files:
                photo = request.files['photo']
                if photo and photo.filename != '':
                    if allowed_file(photo.filename):
                        upload_folder = 'static/uploads/photos'
                        os.makedirs(upload_folder, exist_ok=True)
                        filename = secure_filename(photo.filename)
                        unique_filename = f"{employee_id}_{filename}"
                        photo_path = os.path.join(upload_folder, unique_filename)
                        photo.save(photo_path)
                        photo_path = f"uploads/photos/{unique_filename}"
                    else:
                        flash('❌ Invalid photo format. Please use JPG, PNG, or JPEG files.', 'error')
                        return render_template('edit_faculty.html', faculty=request.form)

            # Handle document upload with validation
            document_path = None
            if 'name_change_proof' in request.files:
                document = request.files['name_change_proof']
                if document and document.filename != '':
                    if allowed_file(document.filename) and document.content_length <= MAX_FILE_SIZE:
                        doc_upload_folder = 'static/uploads/documents'
                        os.makedirs(doc_upload_folder, exist_ok=True)
                        doc_filename = secure_filename(document.filename)
                        unique_docname = f"{employee_id}_proof_{doc_filename}"
                        doc_save_path = os.path.join(doc_upload_folder, unique_docname)
                        document.save(doc_save_path)
                        document_path = f"uploads/documents/{unique_docname}"
                    elif document.content_length > MAX_FILE_SIZE:
                        flash('❌ File size too large. Maximum 5MB allowed.', 'error')
                        return render_template('edit_faculty.html', faculty=request.form)
                    else:
                        flash('❌ Invalid document format. Please use PDF, DOC, DOCX, JPG, or PNG files.', 'error')
                        return render_template('edit_faculty.html', faculty=request.form)

            # Build UPDATE query dynamically based on what fields are provided
            conn = get_db_connection()
            cursor = conn.cursor()
            
            # Basic update query
            update_query = '''
                UPDATE faculty SET 
                employee_id=%s, name_ssc=%s, name_change=%s, dob=%s, gender=%s, 
                blood_group=%s, marital_status=%s, father_name=%s, present_address=%s, 
                permanent_address=%s, email=%s, mobile_no=%s, alternative_mobile=%s, 
                department=%s, designation=%s, date_of_joining=%s, appointment_type=%s, 
                aadhaar_number=%s, pan_number=%s, bank_name=%s, bank_account_no=%s, 
                ifsc_code=%s, caste=%s, subcaste=%s, ratified=%s, ratified_designation=%s,
                ratification_date=%s, previous_employment_date=%s, resignation_date=%s,
                teaching_exp_pragati=%s, teaching_exp_other=%s, industrial_exp=%s, 
                overall_exp=%s, experience_category=%s
            '''
            
            params = [
                employee_id, name_ssc, name_change, dob, gender, blood_group, marital_status,
                father_name, present_address, permanent_address, email, mobile_no, alternative_mobile,
                department, designation, date_of_joining, appointment_type, aadhaar_number, pan_number,
                bank_name, bank_account_no, ifsc_code, caste, subcaste, ratified, ratified_designation,
                ratification_date, previous_employment_date, resignation_date, teaching_exp_pragati,
                teaching_exp_other, industrial_exp, overall_exp, experience_category
            ]
            
            # Add photo path if new photo uploaded
            if photo_path:
                update_query += ', photo_path=%s'
                params.append(photo_path)
            
            # Add document path if new document uploaded
            if document_path:
                update_query += ', name_change_proof=%s'
                params.append(document_path)
            
            # Handle photo removal
            elif 'remove_photo' in request.form:
                update_query += ', photo_path=NULL'
            
            # Handle document removal  
            elif 'remove_name_change_proof' in request.form:
                update_query += ', name_change_proof=NULL'
            
            # Add WHERE clause
            update_query += ' WHERE id=%s'
            params.append(faculty_id)
            
            print(f"DEBUG: Executing update query for faculty_id: {faculty_id}")
            cursor.execute(update_query, params)
            conn.commit()
            cursor.close()
            conn.close()
            
            flash('✅ Faculty information updated successfully!', 'success')
            return redirect('/faculty')
            
        except mysql.connector.Error as err:
            # Handle specific database errors
            if 'conn' in locals() and conn.is_connected():
                conn.rollback()
                cursor.close()
                conn.close()
            
            error_messages = {
                1062: "❌ Duplicate entry detected. Employee ID or Email already exists.",
                1452: "❌ Reference error. Please check department or other related data.",
                1406: "❌ Data too long for one or more fields.",
                1366: "❌ Incorrect data format in one or more fields.",
                1048: "❌ Required field is missing. Please check all mandatory fields."
            }
            
            user_message = error_messages.get(err.errno, f'❌ Database error: {str(err)}')
            flash(user_message, 'error')
            return render_template('edit_faculty.html', faculty=request.form)
            
        except Exception as e:
            print(f"DEBUG: Error in edit_faculty: {str(e)}")
            flash(f'❌ Error updating faculty: {str(e)}', 'error')
            return render_template('edit_faculty.html', faculty=request.form)
    
    # GET request - load existing data
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute('SELECT * FROM faculty WHERE id = %s', (faculty_id,))
        faculty = cursor.fetchone()
        cursor.close()
        conn.close()
        
        if not faculty:
            flash('❌ Faculty member not found!', 'error')
            return redirect('/faculty')
        
        return render_template('edit_faculty.html', faculty=faculty)
        
    except Exception as e:
        flash(f'❌ Error loading faculty data: {str(e)}', 'error')
        return redirect('/faculty')

@app.route('/delete_faculty/<int:faculty_id>')
@login_required
def delete_faculty(faculty_id):
    try:
        # First check if faculty exists
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute('SELECT name_ssc FROM faculty WHERE id = %s', (faculty_id,))
        faculty = cursor.fetchone()
        
        if not faculty:
            flash('❌ Faculty member not found!', 'error')
            return redirect('/faculty')
        
        # Delete the faculty member
        cursor.execute('DELETE FROM faculty WHERE id = %s', (faculty_id,))
        conn.commit()
        cursor.close()
        conn.close()
        
        flash(f'✅ Faculty member {faculty["name_ssc"]} deleted successfully!', 'success')
        return redirect('/faculty')
        
    except mysql.connector.Error as err:
        if 'conn' in locals() and conn.is_connected():
            conn.rollback()
            cursor.close()
            conn.close()
        
        if err.errno == 1451:  # Foreign key constraint violation
            flash('❌ Cannot delete faculty member. This faculty has related records (qualifications, etc.). Please delete related records first.', 'error')
        else:
            flash(f'❌ Database error while deleting faculty: {str(err)}', 'error')
        return redirect('/faculty')
        
    except Exception as e:
        flash(f'❌ Unexpected error while deleting faculty: {str(e)}', 'error')
        return redirect('/faculty')

@app.route('/faculty/<int:faculty_id>/qualifications')
@login_required
def view_qualifications(faculty_id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    
    # Get faculty details
    cursor.execute('SELECT * FROM faculty WHERE id = %s', (faculty_id,))
    faculty = cursor.fetchone()
    
    # Get qualifications
    cursor.execute('SELECT * FROM qualifications WHERE faculty_id = %s ORDER BY year_of_passing DESC', (faculty_id,))
    qualifications = cursor.fetchall()
    
    cursor.close()
    conn.close()
    
    return render_template('qualifications.html', faculty=faculty, qualifications=qualifications)

@app.route('/add_qualification/<int:faculty_id>', methods=['POST'])
@login_required
def add_qualification(faculty_id):
    qualification_type = request.form['qualification_type']
    domain_specialization = request.form.get('domain_specialization', '')
    percentage = request.form.get('percentage', '')
    year_of_passing = request.form.get('year_of_passing', '')
    institution_name = request.form['institution_name']
    highest_degree = 'highest_degree' in request.form
    pursuing = 'pursuing' in request.form
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    cursor.execute(
        '''INSERT INTO qualifications 
        (faculty_id, qualification_type, domain_specialization, percentage, year_of_passing, institution_name, highest_degree, pursuing) 
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s)''',
        (faculty_id, qualification_type, domain_specialization, percentage, year_of_passing, institution_name, highest_degree, pursuing)
    )
    
    conn.commit()
    cursor.close()
    conn.close()
    flash('✅ Qualification added successfully!', 'success')
    return redirect(f'/faculty/{faculty_id}/qualifications')

@app.route('/delete_qualification/<int:qualification_id>')
@login_required
def delete_qualification(qualification_id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    
    # Get faculty_id before deleting
    cursor.execute('SELECT faculty_id FROM qualifications WHERE id = %s', (qualification_id,))
    qualification = cursor.fetchone()
    faculty_id = qualification['faculty_id']
    
    # Delete qualification
    cursor.execute('DELETE FROM qualifications WHERE id = %s', (qualification_id,))
    
    conn.commit()
    cursor.close()
    conn.close()
    flash('✅ Qualification deleted successfully!', 'success')
    return redirect(f'/faculty/{faculty_id}/qualifications')

@app.route('/faculty/<int:faculty_id>')
@login_required
def view_faculty(faculty_id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    
    # Get faculty details
    cursor.execute('SELECT * FROM faculty WHERE id = %s', (faculty_id,))
    faculty = cursor.fetchone()
    
    # 🔒 ACCESS CONTROL: Faculty can only view their own profile
    if get_user_role() in ['viewer'] and faculty['email'] != session.get('email'):
        flash('❌ Access denied. You can only view your own profile.', 'error')
        return redirect('/faculty')
    
    # Get qualifications
    cursor.execute('SELECT * FROM qualifications WHERE faculty_id = %s ORDER BY year_of_passing DESC', (faculty_id,))
    qualifications = cursor.fetchall()
    
    cursor.close()
    conn.close()
    
    return render_template('view_faculty.html', faculty=faculty, qualifications=qualifications)

@app.route('/department/<department_name>')
@login_required
def department_details(department_name):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    
    # 🔒 FOR FACULTY USERS: Check if they have a profile in this department
    if get_user_role() == 'Faculty':
        # First, check if faculty exists in this department with their email
        cursor.execute('SELECT * FROM faculty WHERE department = %s AND email = %s', 
                      (department_name, session.get('email')))
        faculty_in_dept = cursor.fetchone()
        
        if not faculty_in_dept:
            # Faculty doesn't belong to this department - redirect to their profile
            flash(f'❌ Access Denied: You do not have a profile in the {department_name} department.', 'error')
            
            # Find their actual department
            cursor.execute('SELECT department FROM faculty WHERE email = %s', (session.get('email'),))
            actual_faculty = cursor.fetchone()
            
            cursor.close()
            conn.close()
            
            if actual_faculty:
                # Redirect to their actual department
                return redirect(f'/department/{actual_faculty["department"]}')
            else:
                # No profile exists at all
                return redirect('/faculty')
    
    # 🔒 ROLE-BASED DATA ACCESS
    if get_user_role() == 'Faculty':
        # Faculty can only see their own data in this department
        cursor.execute('''
            SELECT * FROM faculty 
            WHERE department = %s AND email = %s
            ORDER BY name_ssc
        ''', (department_name, session.get('email')))
    else:
        # IQAC/Office/Admin see all faculty in this department
        cursor.execute('''
            SELECT * FROM faculty 
            WHERE department = %s 
            ORDER BY name_ssc
        ''', (department_name,))
    
    faculty = cursor.fetchall()
    
    # Get department statistics
    cursor.execute('''
        SELECT 
            COUNT(*) as total,
            COUNT(CASE WHEN gender = 'M' THEN 1 END) as male_count,
            COUNT(CASE WHEN gender = 'F' THEN 1 END) as female_count,
            COUNT(CASE WHEN ratified = 'Yes' THEN 1 END) as ratified_count
        FROM faculty 
        WHERE department = %s
    ''', (department_name,))
    
    stats = cursor.fetchone()
    
    cursor.close()
    conn.close()
    
    return render_template('department_details.html', 
                         department_name=department_name,
                         faculty=faculty,
                         stats=stats)

@app.route('/experience/<experience_category>')
@login_required
def experience_details(experience_category):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    
    # 🔒 FOR FACULTY USERS: Check if they belong to this experience category
    if get_user_role() == 'Faculty':
        cursor.execute('SELECT * FROM faculty WHERE experience_category = %s AND email = %s', 
                      (experience_category, session.get('email')))
        faculty_in_category = cursor.fetchone()
        
        if not faculty_in_category:
            flash(f'❌ Access Denied: You do not belong to the {experience_category} years experience category.', 'error')
            
            # Find their actual experience category
            cursor.execute('SELECT experience_category FROM faculty WHERE email = %s', (session.get('email'),))
            actual_faculty = cursor.fetchone()
            
            cursor.close()
            conn.close()
            
            if actual_faculty:
                return redirect(f'/experience/{actual_faculty["experience_category"]}')
            else:
                return redirect('/faculty')
    
    # 🔒 ROLE-BASED ACCESS: Viewers can see experience pages
    if get_user_role() in ['viewer']:
        cursor.execute('''
            SELECT * FROM faculty 
            WHERE experience_category = %s 
            ORDER BY department, name_ssc
        ''', (experience_category,))
    else:
        cursor.execute('''
            SELECT * FROM faculty 
            WHERE experience_category = %s 
            ORDER BY 
                CASE 
                    WHEN designation = 'Professor' THEN 1
                    WHEN designation = 'Associate Professor' THEN 2
                    WHEN designation = 'Assistant Professor' THEN 3
                    ELSE 4
                END,
                ratified ASC,
                department,
                name_ssc
        ''', (experience_category,))
    
    faculty = cursor.fetchall()
    
    # Get experience statistics (viewers can see stats)
    cursor.execute('''
        SELECT 
            COUNT(*) as total,
            COUNT(CASE WHEN gender = 'M' THEN 1 END) as male_count,
            COUNT(CASE WHEN gender = 'F' THEN 1 END) as female_count,
            COUNT(CASE WHEN ratified = 'Yes' THEN 1 END) as ratified_count,
            COUNT(DISTINCT department) as department_count
        FROM faculty 
        WHERE experience_category = %s
    ''', (experience_category,))
    
    stats = cursor.fetchone()
    
    cursor.close()
    conn.close()
    
    return render_template('experience_details.html', 
                         experience_category=experience_category,
                         faculty=faculty,
                         stats=stats)

@app.route('/designation/<designation_name>')
@login_required
def designation_details(designation_name):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    
    # 🔒 FOR FACULTY USERS: Check if they have this designation
    if get_user_role() == 'Faculty':
        cursor.execute('SELECT * FROM faculty WHERE designation = %s AND email = %s', 
                      (designation_name, session.get('email')))
        faculty_with_designation = cursor.fetchone()
        
        if not faculty_with_designation:
            flash(f'❌ Access Denied: You do not have the {designation_name} designation.', 'error')
            
            # Find their actual designation
            cursor.execute('SELECT designation FROM faculty WHERE email = %s', (session.get('email'),))
            actual_faculty = cursor.fetchone()
            
            cursor.close()
            conn.close()
            
            if actual_faculty:
                return redirect(f'/designation/{actual_faculty["designation"]}')
            else:
                return redirect('/faculty')
    
    # 🔒 ROLE-BASED ACCESS: Viewers can see designation pages
    if get_user_role() in ['viewer']:
        cursor.execute('''
            SELECT * FROM faculty 
            WHERE designation = %s 
            ORDER BY department, name_ssc
        ''', (designation_name,))
    else:
        cursor.execute('''
            SELECT * FROM faculty 
            WHERE designation = %s 
            ORDER BY 
                department,
                CASE experience_category
                    WHEN '10+' THEN 1
                    WHEN '6-10' THEN 2
                    WHEN '0-5' THEN 3
                    ELSE 4
                END,
                name_ssc
        ''', (designation_name,))
    
    faculty = cursor.fetchall()
    
    # Get designation statistics (viewers can see stats)
    cursor.execute('''
        SELECT 
            department,
            COUNT(*) as total,
            COUNT(CASE WHEN gender = 'M' THEN 1 END) as male_count,
            COUNT(CASE WHEN gender = 'F' THEN 1 END) as female_count,
            COUNT(CASE WHEN ratified = 'Yes' THEN 1 END) as ratified_count,
            COUNT(CASE WHEN appointment_type = 'Regular' THEN 1 END) as regular_count
        FROM faculty 
        WHERE designation = %s
        GROUP BY department
        ORDER BY total DESC
    ''', (designation_name,))
    
    department_stats = cursor.fetchall()
    
    # Get overall designation statistics
    cursor.execute('''
        SELECT 
            COUNT(*) as total,
            COUNT(CASE WHEN gender = 'M' THEN 1 END) as male_count,
            COUNT(CASE WHEN gender = 'F' THEN 1 END) as female_count,
            COUNT(CASE WHEN ratified = 'Yes' THEN 1 END) as ratified_count,
            COUNT(CASE WHEN appointment_type = 'Regular' THEN 1 END) as regular_count,
            COUNT(DISTINCT department) as department_count
        FROM faculty 
        WHERE designation = %s
    ''', (designation_name,))
    
    stats = cursor.fetchone()
    
    cursor.close()
    conn.close()
    
    return render_template('designation_details.html', 
                         designation_name=designation_name,
                         faculty=faculty,
                         department_stats=department_stats,
                         stats=stats)

@app.route('/manage_users')
@login_required
def manage_users():
    # Only IQAC can manage users
    if get_user_role() != 'IQAC':
        flash('❌ Access denied. IQAC privileges required.', 'error')
        return redirect('/')
    
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute('SELECT * FROM users ORDER BY approved ASC, role, username')
    users = cursor.fetchall()
    
    # Get pending users count
    cursor.execute('SELECT COUNT(*) as pending_count FROM users WHERE approved = FALSE')
    pending_count = cursor.fetchone()['pending_count']
    
    cursor.close()
    conn.close()
    
    return render_template('manage_users.html', users=users, pending_count=pending_count)

@app.route('/delete_user/<int:user_id>')
@login_required
def delete_user(user_id):
    # Only IQAC can delete users
    if get_user_role() != 'IQAC':
        flash('❌ Access denied. IQAC privileges required.', 'error')
        return redirect('/')
    
    # Prevent IQAC from deleting themselves
    if user_id == session['user_id']:
        flash('❌ You cannot delete your own account!', 'error')
        return redirect('/manage_users')
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        # First, get the username for the flash message
        cursor.execute('SELECT username FROM users WHERE id = %s', (user_id,))
        user_to_delete = cursor.fetchone()
        
        if user_to_delete:
            cursor.execute('DELETE FROM users WHERE id = %s', (user_id,))
            conn.commit()
            flash(f'✅ User {user_to_delete[0]} deleted successfully!', 'success')
        else:
            flash('❌ User not found!', 'error')
            
    except Exception as e:
        flash(f'❌ Error deleting user: {str(e)}', 'error')
    
    finally:
        cursor.close()
        conn.close()
    
    return redirect('/manage_users')

# =====================
# USER APPROVAL ROUTES
# =====================

@app.route('/approve_users')
@login_required
def approve_users():
    # Only IQAC can manage users
    if get_user_role() != 'IQAC':
        flash('❌ Access denied. IQAC privileges required.', 'error')
        return redirect('/')
    
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute('SELECT * FROM users WHERE approved = FALSE ORDER BY created_at DESC')
    pending_users = cursor.fetchall()
    cursor.close()
    conn.close()
    
    return render_template('approve_users.html', pending_users=pending_users)

@app.route('/approve_user/<int:user_id>')
@login_required
def approve_user(user_id):
    # Only IQAC can approve users
    if get_user_role() != 'IQAC':
        flash('❌ IQAC access required', 'error')
        return redirect('/')
    
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('UPDATE users SET approved = TRUE WHERE id = %s', (user_id,))
    conn.commit()
    cursor.close()
    conn.close()
    
    flash('✅ User approved successfully!', 'success')
    return redirect('/approve_users')

@app.route('/reject_user/<int:user_id>')
@login_required
def reject_user(user_id):
    # Only IQAC can reject users
    if get_user_role() != 'IQAC':
        flash('❌ IQAC access required', 'error')
        return redirect('/')
    
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('DELETE FROM users WHERE id = %s AND approved = FALSE', (user_id,))
    conn.commit()
    cursor.close()
    conn.close()
    
    flash('✅ User registration rejected!', 'success')
    return redirect('/approve_users')

@app.route('/download_faculty_excel')
@login_required
def download_faculty_excel():
    try:
        # DEBUG: Log all incoming parameters
        print("🎯 EXCEL DOWNLOAD DEBUG START =========")
        print(f"Full request URL: {request.url}")
        print(f"All request args: {dict(request.args)}")
        
        # Get ALL filter parameters - CORRECTED VARIABLE NAMES
        search = request.args.get('search', '')
        department = request.args.get('department', '')
        appointment_type = request.args.get('appointment_type', '')
        exp_from = request.args.get('exp_from', '')  # CORRECT: exp_from
        exp_to = request.args.get('exp_to', '')      # CORRECT: exp_to
        designation = request.args.get('designation', '')

        print(f"🎯 EXCEL DOWNLOAD FILTERS:")
        print(f"   - search: '{search}'")
        print(f"   - department: '{department}'") 
        print(f"   - designation: '{designation}'")
        print(f"   - appointment_type: '{appointment_type}'")
        print(f"   - exp_from: '{exp_from}'")
        print(f"   - exp_to: '{exp_to}'")
        
        # 🔒 ROLE-BASED DATA ACCESS
        if get_user_role() in ['Faculty']:
            query = 'SELECT * FROM faculty WHERE email = %s'
            params = [session.get('email')]
        else:
            query = 'SELECT * FROM faculty WHERE 1=1'
            params = []
        
        # Apply filters
        if search:
            query += ' AND (name_ssc LIKE %s OR employee_id LIKE %s)'
            params.extend([f'%{search}%', f'%{search}%'])
        
        if department:
            query += ' AND department = %s'
            params.append(department)
        
        if designation:
            query += ' AND designation = %s'
            params.append(designation)
        
        if appointment_type:
            query += ' AND appointment_type = %s'
            params.append(appointment_type)
        
        # CORRECTED: Experience range filter
        if exp_from and exp_from.strip():
            try:
                query += ' AND overall_exp >= %s'
                params.append(float(exp_from))
                print(f"🔍 DEBUG: Added exp_from filter: overall_exp >= {exp_from}")
            except ValueError:
                print(f"⚠️ Invalid exp_from value in Excel: {exp_from}")

        if exp_to and exp_to.strip():
            try:
                query += ' AND overall_exp <= %s'
                params.append(float(exp_to))
                print(f"🔍 DEBUG: Added exp_to filter: overall_exp <= {exp_to}")
            except ValueError:
                print(f"⚠️ Invalid exp_to value in Excel: {exp_to}")
        
        query += ' ORDER BY name_ssc'
        
        print(f"📊 EXCEL QUERY: {query}")
        print(f"🔧 EXCEL PARAMS: {params}")

        # Fetch data
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute(query, params)
        faculty_data = cursor.fetchall()
        
        # NEW: Fetch qualifications for all faculty
        faculty_ids = [str(f['id']) for f in faculty_data]
        qualifications_data = {}
        if faculty_ids:
            placeholders = ','.join(['%s'] * len(faculty_ids))
            cursor.execute(f'''
                SELECT q.*, f.name_ssc, f.employee_id, f.department, f.designation
                FROM qualifications q 
                JOIN faculty f ON q.faculty_id = f.id 
                WHERE q.faculty_id IN ({placeholders})
                ORDER BY q.faculty_id, q.year_of_passing DESC
            ''', faculty_ids)
            all_qualifications = cursor.fetchall()
            
            # Organize qualifications by faculty_id
            for qual in all_qualifications:
                faculty_id = qual['faculty_id']
                if faculty_id not in qualifications_data:
                    qualifications_data[faculty_id] = []
                qualifications_data[faculty_id].append(qual)
        
        cursor.close()
        conn.close()
        
        print(f"✅ EXCEL FOUND {len(faculty_data)} RECORDS")
        if faculty_data:
            for f in faculty_data:
                print(f"   - {f['employee_id']}: {f['name_ssc']} | Dept: {f['department']} | Designation: {f['designation']} | Exp Cat: {f.get('experience_category', 'N/A')}")
        
        # Create Excel workbook with multiple sheets
        wb = openpyxl.Workbook()
        
        # Sheet 1: Faculty Basic Info (your existing sheet)
        ws_faculty = wb.active
        ws_faculty.title = "Faculty Basic Info"

        # Define headers - UPDATED WITH ALTERNATE MOBILE
        headers = [
            'S.No', 'Employee ID', 'Full Name', 'Department', 'Designation', 
            'Total Exp', 'Pragati Exp', 'Appointment Type', 'Email', 'Mobile No',
            'Alternate Mobile', 'Date of Joining', 'Gender', 'Caste', 'Ratified', 'Experience Category'
        ]

        current_row = 1

        # Add filter info as header if any filters are active
        if any([search, department, appointment_type, exp_from, exp_to, designation]):  # CORRECTED: exp_from, exp_to
            ws_faculty.merge_cells(f'A{current_row}:P{current_row}')
            filter_info = "📊 FACULTY DATA EXPORT - Filtered Results: "
            filters = []
            if search: filters.append(f"Search: '{search}'")
            if department: filters.append(f"Department: {department}")
            if designation: filters.append(f"Designation: {designation}")
            if appointment_type: filters.append(f"Appointment: {appointment_type}")
            if exp_from or exp_to:  # CORRECTED: exp_from, exp_to
                exp_filter = f"Experience: {exp_from if exp_from else '0'} to {exp_to if exp_to else '50'} years"
                filters.append(exp_filter)
            
            filter_info += " | ".join(filters)
            ws_faculty.cell(row=current_row, column=1, value=filter_info)
            ws_faculty.cell(row=current_row, column=1).font = Font(bold=True, color="2E86C1", size=12)
            ws_faculty.cell(row=current_row, column=1).alignment = Alignment(horizontal='center')
            current_row += 1

        # Add export info
        ws_faculty.merge_cells(f'A{current_row}:P{current_row}')
        from datetime import datetime
        export_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws_faculty.cell(row=current_row, column=1, value=f"Exported on: {export_time} | Total Records: {len(faculty_data)}")
        ws_faculty.cell(row=current_row, column=1).font = Font(italic=True, color="7D3C98")
        ws_faculty.cell(row=current_row, column=1).alignment = Alignment(horizontal='center')
        current_row += 1

        # Add empty row for spacing
        current_row += 1

        # Add headers with styling
        header_start_row = current_row
        for col, header in enumerate(headers, 1):
            cell = ws_faculty.cell(row=current_row, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal='center')
            cell.fill = openpyxl.styles.PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        
        current_row += 1

        # Add data rows - UPDATED WITH ALTERNATE MOBILE
        if faculty_data:
            for faculty in faculty_data:
                ws_faculty.cell(row=current_row, column=1, value=current_row - header_start_row - 1)  # S.No
                ws_faculty.cell(row=current_row, column=2, value=faculty['employee_id'])
                ws_faculty.cell(row=current_row, column=3, value=faculty['name_ssc'])
                ws_faculty.cell(row=current_row, column=4, value=faculty['department'])
                ws_faculty.cell(row=current_row, column=5, value=faculty['designation'])
                ws_faculty.cell(row=current_row, column=6, value=faculty.get('overall_exp', 0))
                ws_faculty.cell(row=current_row, column=7, value=faculty.get('teaching_exp_pragati', 0))
                ws_faculty.cell(row=current_row, column=8, value=faculty['appointment_type'])
                ws_faculty.cell(row=current_row, column=9, value=faculty['email'])
                ws_faculty.cell(row=current_row, column=10, value=faculty['mobile_no'])
                ws_faculty.cell(row=current_row, column=11, value=faculty.get('alternative_mobile', ''))  # Alternate Mobile
                ws_faculty.cell(row=current_row, column=12, value=str(faculty['date_of_joining']) if faculty['date_of_joining'] else '')
                ws_faculty.cell(row=current_row, column=13, value=faculty['gender'])
                ws_faculty.cell(row=current_row, column=14, value=faculty.get('caste', ''))
                ws_faculty.cell(row=current_row, column=15, value=faculty.get('ratified', 'No'))
                ws_faculty.cell(row=current_row, column=16, value=faculty.get('experience_category', ''))
                current_row += 1
        else:
            # Add "No data" message
            ws_faculty.merge_cells(f'A{current_row}:P{current_row}')
            ws_faculty.cell(row=current_row, column=1, value="❌ NO DATA FOUND - No faculty records match your search criteria")
            ws_faculty.cell(row=current_row, column=1).font = Font(bold=True, color="E74C3C", size=14)
            ws_faculty.cell(row=current_row, column=1).alignment = Alignment(horizontal='center')
            ws_faculty.cell(row=current_row, column=1).fill = openpyxl.styles.PatternFill(start_color="FDEDEC", end_color="FDEDEC", fill_type="solid")

        # Set reasonable column widths - UPDATED
        column_widths = [8, 15, 25, 15, 20, 10, 12, 15, 25, 15, 15, 12, 8, 15, 10, 15]
        for col_idx, width in enumerate(column_widths, 1):
            column_letter = openpyxl.utils.get_column_letter(col_idx)
            ws_faculty.column_dimensions[column_letter].width = width

        # Sheet 2: NEW - Qualifications Sheet
        if qualifications_data:
            ws_qualifications = wb.create_sheet("Qualifications Details")
            
            # Headers for qualifications sheet
            qual_headers = [
                'S.No', 'Employee ID', 'Faculty Name', 'Department', 'Designation',
                'Qualification Type', 'Specialization', 'Institution', 
                'Year of Passing', 'Percentage', 'Highest Degree', 'Pursuing'
            ]
            
            # Add headers
            current_qual_row = 1
            for col, header in enumerate(qual_headers, 1):
                cell = ws_qualifications.cell(row=current_qual_row, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
            
            # Add qualifications data
            current_qual_row = 2
            for faculty in faculty_data:
                faculty_id = faculty['id']
                if faculty_id in qualifications_data:
                    for qual in qualifications_data[faculty_id]:
                        ws_qualifications.cell(row=current_qual_row, column=1, value=current_qual_row-1)
                        ws_qualifications.cell(row=current_qual_row, column=2, value=faculty['employee_id'])
                        ws_qualifications.cell(row=current_qual_row, column=3, value=faculty['name_ssc'])
                        ws_qualifications.cell(row=current_qual_row, column=4, value=faculty['department'])
                        ws_qualifications.cell(row=current_qual_row, column=5, value=faculty['designation'])
                        ws_qualifications.cell(row=current_qual_row, column=6, value=qual['qualification_type'])
                        ws_qualifications.cell(row=current_qual_row, column=7, value=qual.get('domain_specialization', ''))
                        ws_qualifications.cell(row=current_qual_row, column=8, value=qual['institution_name'])
                        ws_qualifications.cell(row=current_qual_row, column=9, value=qual.get('year_of_passing', ''))
                        ws_qualifications.cell(row=current_qual_row, column=10, value=qual.get('percentage', ''))
                        ws_qualifications.cell(row=current_qual_row, column=11, value='Yes' if qual['highest_degree'] else 'No')
                        ws_qualifications.cell(row=current_qual_row, column=12, value='Yes' if qual['pursuing'] else 'No')
                        current_qual_row += 1
            
            # Auto-adjust column widths for qualifications sheet
            for column in ws_qualifications.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min((max_length + 2), 50)
                ws_qualifications.column_dimensions[column_letter].width = adjusted_width

        # Save to bytes buffer
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)

        # Create filename
        from datetime import datetime
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        if faculty_data:
            filename = f"faculty_data_{len(faculty_data)}_records_{timestamp}.xlsx"
        else:
            filename = f"faculty_data_no_results_{timestamp}.xlsx"

        return send_file(
            excel_buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        print(f"❌ ERROR in download_faculty_excel: {str(e)}")
        import traceback
        print(f"🔧 TRACEBACK: {traceback.format_exc()}")
        flash(f'❌ Error generating Excel file: {str(e)}', 'error')
        return redirect('/faculty')

@app.route('/download_faculty_single/<int:faculty_id>')
@login_required
def download_faculty_single(faculty_id):
    try:
        # Fetch single faculty data
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute('SELECT * FROM faculty WHERE id = %s', (faculty_id,))
        faculty = cursor.fetchone()
        
        if not faculty:
            flash('❌ Faculty member not found!', 'error')
            return redirect('/faculty')
        
        # NEW: Fetch qualifications for this faculty
        cursor.execute('SELECT * FROM qualifications WHERE faculty_id = %s ORDER BY year_of_passing DESC', (faculty_id,))
        qualifications = cursor.fetchall()
        
        cursor.close()
        conn.close()
        
        # Create Excel workbook with multiple sheets
        wb = openpyxl.Workbook()
        
        # Sheet 1: Faculty Basic Info
        ws_basic = wb.active
        ws_basic.title = "Basic Information"
        
        # Add COMPREHENSIVE faculty details - UPDATED
        details = [
            ['Employee ID', faculty['employee_id']],
            ['Full Name', faculty['name_ssc']],
            ['Name Change', 'Yes' if faculty.get('name_change') else 'No'],
            ['Date of Birth', str(faculty['dob']) if faculty['dob'] else ''],
            ['Gender', faculty['gender']],
            ['Blood Group', faculty.get('blood_group', 'Not specified')],
            ['Marital Status', faculty.get('marital_status', 'Not specified')],
            ['Father Name', faculty['father_name']],
            ['Present Address', faculty['present_address']],
            ['Permanent Address', faculty['permanent_address']],
            ['Email', faculty['email']],
            ['Mobile No', faculty['mobile_no']],
            ['Alternate Mobile', faculty.get('alternative_mobile', 'Not specified')],  # CHANGED HERE
            ['Department', faculty['department']],
            ['Designation', faculty['designation']],
            ['Date of Joining', str(faculty['date_of_joining']) if faculty['date_of_joining'] else ''],
            ['Appointment Type', faculty['appointment_type']],
            ['Aadhaar Number', faculty.get('aadhaar_number', 'Not specified')],
            ['PAN Number', faculty.get('pan_number', 'Not specified')],
            ['Bank Name', faculty.get('bank_name', 'Not specified')],
            ['Bank Account No', faculty.get('bank_account_no', 'Not specified')],
            ['IFSC Code', faculty.get('ifsc_code', 'Not specified')],
            ['Caste', faculty.get('caste', 'Not specified')],
            ['Subcaste', faculty.get('subcaste', 'Not specified')],
            ['Teaching Experience at Pragati', f"{faculty.get('teaching_exp_pragati', 0)} years"],
            ['Teaching Experience at Other Institutions', f"{faculty.get('teaching_exp_other', 0)} years"],
            ['Industrial Experience', f"{faculty.get('industrial_exp', 0)} years"],
            ['Total Experience', f"{faculty.get('overall_exp', 0)} years"],
            ['Experience Category', faculty.get('experience_category', 'Not specified')],
            ['Ratified', faculty.get('ratified', 'No')],
            ['Ratified Designation', faculty.get('ratified_designation', 'Not specified')],
            ['Date of Ratification', str(faculty.get('ratification_date', '')) if faculty.get('ratification_date') else 'Not specified'],
            ['Last Working Date (Previous Employment)', str(faculty.get('previous_employment_date', '')) if faculty.get('previous_employment_date') else 'Not specified'],
            ['Date of Resignation (Pragati)', str(faculty.get('resignation_date', '')) if faculty.get('resignation_date') else 'Not applicable'],
        ]
        
        # Add headers
        ws_basic['A1'] = 'Field'
        ws_basic['B1'] = 'Value'
        ws_basic['A1'].font = Font(bold=True)
        ws_basic['B1'].font = Font(bold=True)
        
        # Add data
        for row, (field, value) in enumerate(details, 2):
            ws_basic.cell(row=row, column=1, value=field)
            ws_basic.cell(row=row, column=2, value=value)
        
        # Auto-adjust column widths for basic info
        for column in ws_basic.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws_basic.column_dimensions[column_letter].width = adjusted_width
        
        # Sheet 2: NEW - Qualifications Sheet
        if qualifications:
            ws_qualifications = wb.create_sheet("Qualifications")
            
            # Headers
            qual_headers = ['S.No', 'Qualification Type', 'Specialization', 'Institution', 
                           'Year of Passing', 'Percentage', 'Highest Degree', 'Pursuing']
            
            # Add headers
            for col, header in enumerate(qual_headers, 1):
                cell = ws_qualifications.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
            
            # Add qualifications data
            for row, qual in enumerate(qualifications, 2):
                ws_qualifications.cell(row=row, column=1, value=row-1)
                ws_qualifications.cell(row=row, column=2, value=qual['qualification_type'])
                ws_qualifications.cell(row=row, column=3, value=qual.get('domain_specialization', ''))
                ws_qualifications.cell(row=row, column=4, value=qual['institution_name'])
                ws_qualifications.cell(row=row, column=5, value=qual.get('year_of_passing', ''))
                ws_qualifications.cell(row=row, column=6, value=qual.get('percentage', ''))
                ws_qualifications.cell(row=row, column=7, value='Yes' if qual['highest_degree'] else 'No')
                ws_qualifications.cell(row=row, column=8, value='Yes' if qual['pursuing'] else 'No')
            
            # Auto-adjust column widths for qualifications
            for column in ws_qualifications.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min((max_length + 2), 50)
                ws_qualifications.column_dimensions[column_letter].width = adjusted_width
        else:
            # Create empty qualifications sheet with message
            ws_qualifications = wb.create_sheet("Qualifications")
            ws_qualifications['A1'] = "No qualifications found for this faculty member."
        
        # Save to bytes buffer
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        filename = f"{faculty['employee_id']}_{faculty['name_ssc']}_complete_profile.xlsx"
        
        return send_file(
            excel_buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        flash(f'❌ Error generating Excel file: {str(e)}', 'error')
        return redirect('/faculty')
# =====================
# RESEARCH PUBLICATIONS ROUTES
# =====================

@app.route('/faculty/<int:faculty_id>/publications')
@login_required
def view_publications(faculty_id):
    # Access control: Anyone can view, but editing restricted
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    
    # Get faculty details
    cursor.execute('SELECT * FROM faculty WHERE id = %s', (faculty_id,))
    faculty = cursor.fetchone()
    
    if not faculty:
        flash('❌ Faculty member not found!', 'error')
        return redirect('/faculty')
    
    # Get all publications data
    cursor.execute('SELECT * FROM journal_publications WHERE faculty_id = %s ORDER BY year_of_publication DESC', (faculty_id,))
    journals = cursor.fetchall()
    
    cursor.execute('SELECT * FROM conference_publications WHERE faculty_id = %s ORDER BY year_of_publication DESC', (faculty_id,))
    conferences = cursor.fetchall()
    
    cursor.execute('SELECT * FROM book_chapters WHERE faculty_id = %s ORDER BY year_of_publication DESC', (faculty_id,))
    book_chapters = cursor.fetchall()
    
    cursor.execute('SELECT * FROM patents WHERE faculty_id = %s ORDER BY filing_date DESC', (faculty_id,))
    patents = cursor.fetchall()
    
    cursor.close()
    conn.close()
    
    return render_template('publications.html', 
                     faculty=faculty, 
                     journals=journals, 
                     conferences=conferences,
                     book_chapters=book_chapters,
                     patents=patents,
                     can_edit=can_edit_publications(faculty_id))

# Journal Publications Routes
@app.route('/add_journal_publication/<int:faculty_id>', methods=['POST'])
@login_required
def add_journal_publication(faculty_id):
    if not can_edit_publications(faculty_id):
        flash('❌ Access denied. You can only edit your own R&D publications.', 'error')
        return redirect(f'/faculty/{faculty_id}/publications')
    
    try:
        # Get all form data
        department = request.form['department']
        first_author = request.form['first_author']
        corresponding_author = request.form['corresponding_author']
        other_authors = request.form.get('other_authors', '')
        faculty_author_position = request.form['faculty_author_position']
        paper_title_apa = request.form['paper_title_apa']
        journal_name = request.form['journal_name']
        volume_issue = request.form.get('volume_issue', '')
        page_numbers = request.form.get('page_numbers', '')
        issn_number = request.form.get('issn_number', '')
        doi = request.form.get('doi', '')
        year_of_publication = int(request.form['year_of_publication'])
        indexing = request.form.get('indexing', '')
        quartile = request.form.get('quartile', '')
        impact_factor = float(request.form.get('impact_factor', 0))
        journal_link = request.form.get('journal_link', '')
        publisher = request.form.get('publisher', '')
        funding_agency = request.form.get('funding_agency', '')
        remarks = request.form.get('remarks', '')
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute('''
            INSERT INTO journal_publications 
            (faculty_id, department, first_author, corresponding_author, other_authors, 
             faculty_author_position, paper_title_apa, journal_name, volume_issue, 
             page_numbers, issn_number, doi, year_of_publication, indexing, quartile, 
             impact_factor, journal_link, publisher, funding_agency, remarks) 
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        ''', (faculty_id, department, first_author, corresponding_author, other_authors,
              faculty_author_position, paper_title_apa, journal_name, volume_issue,
              page_numbers, issn_number, doi, year_of_publication, indexing, quartile,
              impact_factor, journal_link, publisher, funding_agency, remarks))
        
        conn.commit()
        cursor.close()
        conn.close()
        
        flash('✅ Journal publication added successfully!', 'success')
        return redirect(f'/faculty/{faculty_id}/publications')
        
    except Exception as e:
        flash(f'❌ Error adding journal publication: {str(e)}', 'error')
        return redirect(f'/faculty/{faculty_id}/publications')

@app.route('/delete_journal/<int:journal_id>')
@login_required
def delete_journal(journal_id):
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Get faculty_id before deleting
        cursor.execute('SELECT faculty_id FROM journal_publications WHERE id = %s', (journal_id,))
        journal = cursor.fetchone()
        if not journal:
            flash('❌ Journal publication not found!', 'error')
            return redirect('/faculty')
        
        # Check if user can edit this faculty's publications
        if not can_edit_publications(journal['faculty_id']):
            flash('❌ Access denied. You can only delete your own R&D publications.', 'error')
            return redirect(f'/faculty/{journal["faculty_id"]}/publications')
        if journal:
            cursor.execute('DELETE FROM journal_publications WHERE id = %s', (journal_id,))
            conn.commit()
            flash('✅ Journal publication deleted successfully!', 'success')
        else:
            flash('❌ Journal publication not found!', 'error')
            
        cursor.close()
        conn.close()
        
        return redirect(f'/faculty/{journal["faculty_id"]}/publications')
        
    except Exception as e:
        flash(f'❌ Error deleting journal publication: {str(e)}', 'error')
        return redirect('/faculty')

# Conference Publications Routes
@app.route('/add_conference_publication/<int:faculty_id>', methods=['POST'])
@login_required
def add_conference_publication(faculty_id):
    # Check if user can edit this faculty's publications
    if not can_edit_publications(faculty_id):
        flash('❌ Access denied. You can only edit your own R&D publications.', 'error')
        return redirect(f'/faculty/{faculty_id}/publications')
    
    try:
        department = request.form['department']
        paper_title = request.form['paper_title']
        authors = request.form['authors']
        corresponding_author = request.form['corresponding_author']
        faculty_author_position = request.form['faculty_author_position']
        conference_name = request.form['conference_name']
        conference_venue = request.form.get('conference_venue', '')
        conference_dates = request.form.get('conference_dates', '')
        proceedings_title = request.form.get('proceedings_title', '')
        isbn_issn = request.form.get('isbn_issn', '')
        doi = request.form.get('doi', '')
        year_of_publication = int(request.form['year_of_publication'])
        indexing = request.form.get('indexing', '')
        publisher = request.form.get('publisher', '')
        conference_link = request.form.get('conference_link', '')
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute('''
            INSERT INTO conference_publications 
            (faculty_id, department, paper_title, authors, corresponding_author, 
             faculty_author_position, conference_name, conference_venue, conference_dates,
             proceedings_title, isbn_issn, doi, year_of_publication, indexing, publisher, conference_link) 
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        ''', (faculty_id, department, paper_title, authors, corresponding_author,
              faculty_author_position, conference_name, conference_venue, conference_dates,
              proceedings_title, isbn_issn, doi, year_of_publication, indexing, publisher, conference_link))
        
        conn.commit()
        cursor.close()
        conn.close()
        
        flash('✅ Conference publication added successfully!', 'success')
        return redirect(f'/faculty/{faculty_id}/publications')
        
    except Exception as e:
        flash(f'❌ Error adding conference publication: {str(e)}', 'error')
        return redirect(f'/faculty/{faculty_id}/publications')

@app.route('/delete_conference/<int:conference_id>')
@login_required
def delete_conference(conference_id):
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Get faculty_id before deleting
        cursor.execute('SELECT faculty_id FROM conference_publications WHERE id = %s', (conference_id,))
        conference = cursor.fetchone()
        
        if not conference:
            flash('❌ Conference publication not found!', 'error')
            return redirect('/faculty')
        
        # Check if user can edit this faculty's publications
        if not can_edit_publications(conference['faculty_id']):
            flash('❌ Access denied. You can only delete your own R&D publications.', 'error')
            return redirect(f'/faculty/{conference["faculty_id"]}/publications')
        
        cursor.execute('DELETE FROM conference_publications WHERE id = %s', (conference_id,))
        conn.commit()
        cursor.close()
        conn.close()
        
        flash('✅ Conference publication deleted successfully!', 'success')
        return redirect(f'/faculty/{conference["faculty_id"]}/publications')
        
    except Exception as e:
        flash(f'❌ Error deleting conference publication: {str(e)}', 'error')
        return redirect('/faculty')

# Book Chapters Routes
@app.route('/add_book_chapter/<int:faculty_id>', methods=['POST'])
@login_required
def add_book_chapter(faculty_id):
    # Check if user can edit this faculty's publications
    if not can_edit_publications(faculty_id):
        flash('❌ Access denied. You can only edit your own R&D publications.', 'error')
        return redirect(f'/faculty/{faculty_id}/publications')
    
    try:
        department = request.form['department']
        chapter_title = request.form['chapter_title']
        book_title = request.form['book_title']
        authors = request.form['authors']
        faculty_author_position = request.form['faculty_author_position']
        corresponding_author = request.form['corresponding_author']
        publisher = request.form['publisher']
        isbn_number = request.form.get('isbn_number', '')
        chapter_doi = request.form.get('chapter_doi', '')
        year_of_publication = int(request.form['year_of_publication'])
        indexing = request.form.get('indexing', '')
        quartile = request.form.get('quartile', '')
        impact_factor = float(request.form.get('impact_factor', 0))
        chapter_link = request.form.get('chapter_link', '')
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute('''
            INSERT INTO book_chapters 
            (faculty_id, department, chapter_title, book_title, authors, 
             faculty_author_position, corresponding_author, publisher, isbn_number,
             chapter_doi, year_of_publication, indexing, quartile, impact_factor, chapter_link) 
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        ''', (faculty_id, department, chapter_title, book_title, authors,
              faculty_author_position, corresponding_author, publisher, isbn_number,
              chapter_doi, year_of_publication, indexing, quartile, impact_factor, chapter_link))
        
        conn.commit()
        cursor.close()
        conn.close()
        
        flash('✅ Book chapter added successfully!', 'success')
        return redirect(f'/faculty/{faculty_id}/publications')
        
    except Exception as e:
        flash(f'❌ Error adding book chapter: {str(e)}', 'error')
        return redirect(f'/faculty/{faculty_id}/publications')

@app.route('/delete_book_chapter/<int:chapter_id>')
@login_required
def delete_book_chapter(chapter_id):
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Get faculty_id before deleting
        cursor.execute('SELECT faculty_id FROM book_chapters WHERE id = %s', (chapter_id,))
        chapter = cursor.fetchone()
        
        if not chapter:
            flash('❌ Book chapter not found!', 'error')
            return redirect('/faculty')
        
        # Check if user can edit this faculty's publications
        if not check_publication_access(chapter['faculty_id']):
            flash('❌ Access denied. You can only delete your own R&D publications.', 'error')
            return redirect(f'/faculty/{chapter["faculty_id"]}/publications')
        
        cursor.execute('DELETE FROM book_chapters WHERE id = %s', (chapter_id,))
        conn.commit()
        cursor.close()
        conn.close()
        
        flash('✅ Book chapter deleted successfully!', 'success')
        return redirect(f'/faculty/{chapter["faculty_id"]}/publications')
        
    except Exception as e:
        flash(f'❌ Error deleting book chapter: {str(e)}', 'error')
        return redirect('/faculty')

# Patents Routes
@app.route('/add_patent/<int:faculty_id>', methods=['POST'])
@login_required
def add_patent(faculty_id):
    # Check if user can edit this faculty's publications
    if not check_publication_access(faculty_id):
        flash('❌ Access denied. You can only edit your own R&D publications.', 'error')
        return redirect(f'/faculty/{faculty_id}/publications')
    
    try:
        department = request.form['department']
        patent_title = request.form['patent_title']
        inventors = request.form['inventors']
        corresponding_applicant = request.form['corresponding_applicant']
        faculty_author_position = request.form['faculty_author_position']
        patent_application_number = request.form['patent_application_number']
        filing_date = request.form.get('filing_date')
        publication_date = request.form.get('publication_date')
        grant_date = request.form.get('grant_date')
        patent_office = request.form['patent_office']
        status = request.form['status']
        patent_type = request.form['patent_type']
        patent_link = request.form.get('patent_link', '')
        certificate_link = request.form.get('certificate_link', '')
        
        # Convert empty dates to None
        filing_date = filing_date if filing_date else None
        publication_date = publication_date if publication_date else None
        grant_date = grant_date if grant_date else None
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute('''
            INSERT INTO patents 
            (faculty_id, department, patent_title, inventors, corresponding_applicant, 
             faculty_author_position, patent_application_number, filing_date, publication_date,
             grant_date, patent_office, status, patent_type, patent_link, certificate_link) 
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        ''', (faculty_id, department, patent_title, inventors, corresponding_applicant,
              faculty_author_position, patent_application_number, filing_date, publication_date,
              grant_date, patent_office, status, patent_type, patent_link, certificate_link))
        
        conn.commit()
        cursor.close()
        conn.close()
        
        flash('✅ Patent added successfully!', 'success')
        return redirect(f'/faculty/{faculty_id}/publications')
        
    except Exception as e:
        flash(f'❌ Error adding patent: {str(e)}', 'error')
        return redirect(f'/faculty/{faculty_id}/publications')

@app.route('/delete_patent/<int:patent_id>')
@login_required
def delete_patent(patent_id):
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Get faculty_id before deleting
        cursor.execute('SELECT faculty_id FROM patents WHERE id = %s', (patent_id,))
        patent = cursor.fetchone()
        
        if not patent:
            flash('❌ Patent not found!', 'error')
            return redirect('/faculty')
        
        # Check if user can edit this faculty's publications
        if not check_publication_access(patent['faculty_id']):
            flash('❌ Access denied. You can only delete your own R&D publications.', 'error')
            return redirect(f'/faculty/{patent["faculty_id"]}/publications')
        
        cursor.execute('DELETE FROM patents WHERE id = %s', (patent_id,))
        conn.commit()
        cursor.close()
        conn.close()
        
        flash('✅ Patent deleted successfully!', 'success')
        return redirect(f'/faculty/{patent["faculty_id"]}/publications')
        
    except Exception as e:
        flash(f'❌ Error deleting patent: {str(e)}', 'error')
        return redirect('/faculty') 

# View Detailed Routes
@app.route('/view_journal/<int:journal_id>')
@login_required
def view_journal(journal_id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    
    cursor.execute('''
        SELECT j.*, f.name_ssc, f.department as faculty_department 
        FROM journal_publications j 
        JOIN faculty f ON j.faculty_id = f.id 
        WHERE j.id = %s
    ''', (journal_id,))
    journal = cursor.fetchone()
    
    cursor.close()
    conn.close()
    
    if not journal:
        flash('❌ Journal publication not found!', 'error')
        return redirect('/faculty')
    
    return render_template('view_journal.html', journal=journal)

@app.route('/view_conference/<int:conference_id>')
@login_required
def view_conference(conference_id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    
    cursor.execute('''
        SELECT c.*, f.name_ssc, f.department as faculty_department 
        FROM conference_publications c 
        JOIN faculty f ON c.faculty_id = f.id 
        WHERE c.id = %s
    ''', (conference_id,))
    conference = cursor.fetchone()
    
    cursor.close()
    conn.close()
    
    if not conference:
        flash('❌ Conference publication not found!', 'error')
        return redirect('/faculty')
    
    return render_template('view_conference.html', conference=conference)

@app.route('/view_book_chapter/<int:chapter_id>')
@login_required
def view_book_chapter(chapter_id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    
    cursor.execute('''
        SELECT b.*, f.name_ssc, f.department as faculty_department 
        FROM book_chapters b 
        JOIN faculty f ON b.faculty_id = f.id 
        WHERE b.id = %s
    ''', (chapter_id,))
    chapter = cursor.fetchone()
    
    cursor.close()
    conn.close()
    
    if not chapter:
        flash('❌ Book chapter not found!', 'error')
        return redirect('/faculty')
    
    return render_template('view_book_chapter.html', chapter=chapter)

@app.route('/view_patent/<int:patent_id>')
@login_required
def view_patent(patent_id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    
    cursor.execute('''
        SELECT p.*, f.name_ssc, f.department as faculty_department 
        FROM patents p 
        JOIN faculty f ON p.faculty_id = f.id 
        WHERE p.id = %s
    ''', (patent_id,))
    patent = cursor.fetchone()
    
    cursor.close()
    conn.close()
    
    if not patent:
        flash('❌ Patent not found!', 'error')
        return redirect('/faculty')
    
    return render_template('view_patent.html', patent=patent)                               

# =====================
# R&D DOWNLOAD ROUTES
# =====================

@app.route('/download_journals/<int:faculty_id>')
@login_required
def download_journals(faculty_id):
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        cursor.execute('''
            SELECT j.*, f.name_ssc as faculty_name, f.department as faculty_department
            FROM journal_publications j 
            JOIN faculty f ON j.faculty_id = f.id 
            WHERE j.faculty_id = %s 
            ORDER BY j.year_of_publication DESC
        ''', (faculty_id,))
        journals = cursor.fetchall()
        
        cursor.close()
        conn.close()
        
        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Journal Publications"
        
        # Headers
        headers = [
            'S.No', 'Paper Title', 'Journal Name', 'First Author', 'Corresponding Author',
            'Other Authors', 'Faculty Position', 'Volume & Issue', 'Page Numbers',
            'ISSN', 'DOI', 'Year', 'Indexing', 'Quartile', 'Impact Factor',
            'Publisher', 'Funding Agency', 'Journal Link', 'Remarks'
        ]
        
        # Add headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Add data
        for row, journal in enumerate(journals, 2):
            ws.cell(row=row, column=1, value=row-1)
            ws.cell(row=row, column=2, value=journal['paper_title_apa'])
            ws.cell(row=row, column=3, value=journal['journal_name'])
            ws.cell(row=row, column=4, value=journal['first_author'])
            ws.cell(row=row, column=5, value=journal['corresponding_author'])
            ws.cell(row=row, column=6, value=journal['other_authors'] or '')
            ws.cell(row=row, column=7, value=journal['faculty_author_position'])
            ws.cell(row=row, column=8, value=journal['volume_issue'] or '')
            ws.cell(row=row, column=9, value=journal['page_numbers'] or '')
            ws.cell(row=row, column=10, value=journal['issn_number'] or '')
            ws.cell(row=row, column=11, value=journal['doi'] or '')
            ws.cell(row=row, column=12, value=journal['year_of_publication'])
            ws.cell(row=row, column=13, value=journal['indexing'] or '')
            ws.cell(row=row, column=14, value=journal['quartile'] or '')
            ws.cell(row=row, column=15, value=journal['impact_factor'] or '')
            ws.cell(row=row, column=16, value=journal['publisher'] or '')
            ws.cell(row=row, column=17, value=journal['funding_agency'] or '')
            ws.cell(row=row, column=18, value=journal['journal_link'] or '')
            ws.cell(row=row, column=19, value=journal['remarks'] or '')
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min((max_length + 2), 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to bytes buffer
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        # Create filename
        from datetime import datetime
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"journal_publications_{timestamp}.xlsx"
        
        return send_file(
            excel_buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        flash(f'❌ Error generating Excel file: {str(e)}', 'error')
        return redirect(f'/faculty/{faculty_id}/publications')

@app.route('/download_conferences/<int:faculty_id>')
@login_required
def download_conferences(faculty_id):
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        cursor.execute('''
            SELECT c.*, f.name_ssc as faculty_name, f.department as faculty_department
            FROM conference_publications c 
            JOIN faculty f ON c.faculty_id = f.id 
            WHERE c.faculty_id = %s 
            ORDER BY c.year_of_publication DESC
        ''', (faculty_id,))
        conferences = cursor.fetchall()
        
        cursor.close()
        conn.close()
        
        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Conference Publications"
        
        # Headers
        headers = [
            'S.No', 'Paper Title', 'Conference Name', 'Authors', 'Corresponding Author',
            'Faculty Position', 'Venue', 'Dates', 'Proceedings Title', 'ISBN/ISSN',
            'DOI', 'Year', 'Indexing', 'Publisher', 'Conference Link'
        ]
        
        # Add headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Add data
        for row, conference in enumerate(conferences, 2):
            ws.cell(row=row, column=1, value=row-1)
            ws.cell(row=row, column=2, value=conference['paper_title'])
            ws.cell(row=row, column=3, value=conference['conference_name'])
            ws.cell(row=row, column=4, value=conference['authors'])
            ws.cell(row=row, column=5, value=conference['corresponding_author'])
            ws.cell(row=row, column=6, value=conference['faculty_author_position'])
            ws.cell(row=row, column=7, value=conference['conference_venue'] or '')
            ws.cell(row=row, column=8, value=conference['conference_dates'] or '')
            ws.cell(row=row, column=9, value=conference['proceedings_title'] or '')
            ws.cell(row=row, column=10, value=conference['isbn_issn'] or '')
            ws.cell(row=row, column=11, value=conference['doi'] or '')
            ws.cell(row=row, column=12, value=conference['year_of_publication'])
            ws.cell(row=row, column=13, value=conference['indexing'] or '')
            ws.cell(row=row, column=14, value=conference['publisher'] or '')
            ws.cell(row=row, column=15, value=conference['conference_link'] or '')
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min((max_length + 2), 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to bytes buffer
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        # Create filename
        from datetime import datetime
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"conference_publications_{timestamp}.xlsx"
        
        return send_file(
            excel_buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        flash(f'❌ Error generating Excel file: {str(e)}', 'error')
        return redirect(f'/faculty/{faculty_id}/publications')

# Add similar routes for book_chapters and patents following the same pattern
@app.route('/download_book_chapters/<int:faculty_id>')
@login_required
def download_book_chapters(faculty_id):
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        cursor.execute('''
            SELECT b.*, f.name_ssc as faculty_name, f.department as faculty_department
            FROM book_chapters b 
            JOIN faculty f ON b.faculty_id = f.id 
            WHERE b.faculty_id = %s 
            ORDER BY b.year_of_publication DESC
        ''', (faculty_id,))
        chapters = cursor.fetchall()
        
        cursor.close()
        conn.close()
        
        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Book Chapters"
        
        # Headers
        headers = [
            'S.No', 'Chapter Title', 'Book Title', 'Authors', 'Corresponding Author',
            'Faculty Position', 'Publisher', 'ISBN', 'Chapter DOI', 'Year',
            'Indexing', 'Quartile', 'Impact Factor', 'Chapter Link'
        ]
        
        # Add headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Add data
        for row, chapter in enumerate(chapters, 2):
            ws.cell(row=row, column=1, value=row-1)
            ws.cell(row=row, column=2, value=chapter['chapter_title'])
            ws.cell(row=row, column=3, value=chapter['book_title'])
            ws.cell(row=row, column=4, value=chapter['authors'])
            ws.cell(row=row, column=5, value=chapter['corresponding_author'])
            ws.cell(row=row, column=6, value=chapter['faculty_author_position'])
            ws.cell(row=row, column=7, value=chapter['publisher'])
            ws.cell(row=row, column=8, value=chapter['isbn_number'] or '')
            ws.cell(row=row, column=9, value=chapter['chapter_doi'] or '')
            ws.cell(row=row, column=10, value=chapter['year_of_publication'])
            ws.cell(row=row, column=11, value=chapter['indexing'] or '')
            ws.cell(row=row, column=12, value=chapter['quartile'] or '')
            ws.cell(row=row, column=13, value=chapter['impact_factor'] or '')
            ws.cell(row=row, column=14, value=chapter['chapter_link'] or '')
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min((max_length + 2), 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to bytes buffer
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        # Create filename
        from datetime import datetime
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"book_chapters_{timestamp}.xlsx"
        
        return send_file(
            excel_buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        flash(f'❌ Error generating Excel file: {str(e)}', 'error')
        return redirect(f'/faculty/{faculty_id}/publications')

@app.route('/download_patents/<int:faculty_id>')
@login_required
def download_patents(faculty_id):
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        cursor.execute('''
            SELECT p.*, f.name_ssc as faculty_name, f.department as faculty_department
            FROM patents p 
            JOIN faculty f ON p.faculty_id = f.id 
            WHERE p.faculty_id = %s 
            ORDER BY p.filing_date DESC
        ''', (faculty_id,))
        patents = cursor.fetchall()
        
        cursor.close()
        conn.close()
        
        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Patents"
        
        # Headers
        headers = [
            'S.No', 'Patent Title', 'Inventors', 'Corresponding Applicant',
            'Faculty Position', 'Application Number', 'Filing Date', 'Publication Date',
            'Grant Date', 'Patent Office', 'Status', 'Patent Type', 'Patent Link',
            'Certificate Link'
        ]
        
        # Add headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Add data
        for row, patent in enumerate(patents, 2):
            ws.cell(row=row, column=1, value=row-1)
            ws.cell(row=row, column=2, value=patent['patent_title'])
            ws.cell(row=row, column=3, value=patent['inventors'])
            ws.cell(row=row, column=4, value=patent['corresponding_applicant'])
            ws.cell(row=row, column=5, value=patent['faculty_author_position'])
            ws.cell(row=row, column=6, value=patent['patent_application_number'])
            ws.cell(row=row, column=7, value=str(patent['filing_date']) if patent['filing_date'] else '')
            ws.cell(row=row, column=8, value=str(patent['publication_date']) if patent['publication_date'] else '')
            ws.cell(row=row, column=9, value=str(patent['grant_date']) if patent['grant_date'] else '')
            ws.cell(row=row, column=10, value=patent['patent_office'])
            ws.cell(row=row, column=11, value=patent['status'])
            ws.cell(row=row, column=12, value=patent['patent_type'])
            ws.cell(row=row, column=13, value=patent['patent_link'] or '')
            ws.cell(row=row, column=14, value=patent['certificate_link'] or '')
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min((max_length + 2), 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to bytes buffer
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        # Create filename
        from datetime import datetime
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"patents_{timestamp}.xlsx"
        
        return send_file(
            excel_buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        flash(f'❌ Error generating Excel file: {str(e)}', 'error')
        return redirect(f'/faculty/{faculty_id}/publications')

# =====================
# R&D EDIT ROUTES
# =====================

# Edit Journal Publication
@app.route('/edit_journal/<int:journal_id>', methods=['GET', 'POST'])
@login_required
def edit_journal(journal_id):
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Get journal details with faculty info
        cursor.execute('''
            SELECT j.*, f.id as faculty_id, f.name_ssc, f.email 
            FROM journal_publications j 
            JOIN faculty f ON j.faculty_id = f.id 
            WHERE j.id = %s
        ''', (journal_id,))
        journal = cursor.fetchone()
        
        if not journal:
            flash('❌ Journal publication not found!', 'error')
            return redirect('/faculty')
        
        # Check if user can edit this faculty's publications
        if not can_edit_publications(journal['faculty_id']):
            flash('❌ Access denied. You can only edit your own R&D publications.', 'error')
            return redirect(f'/faculty/{journal["faculty_id"]}/publications')
        
        if request.method == 'POST':
            # Get form data
            department = request.form['department']
            first_author = request.form['first_author']
            corresponding_author = request.form['corresponding_author']
            other_authors = request.form.get('other_authors', '')
            faculty_author_position = request.form['faculty_author_position']
            paper_title_apa = request.form['paper_title_apa']
            journal_name = request.form['journal_name']
            volume_issue = request.form.get('volume_issue', '')
            page_numbers = request.form.get('page_numbers', '')
            issn_number = request.form.get('issn_number', '')
            doi = request.form.get('doi', '')
            year_of_publication = int(request.form['year_of_publication'])
            indexing = request.form.get('indexing', '')
            quartile = request.form.get('quartile', '')
            impact_factor = float(request.form.get('impact_factor', 0))
            journal_link = request.form.get('journal_link', '')
            publisher = request.form.get('publisher', '')
            funding_agency = request.form.get('funding_agency', '')
            remarks = request.form.get('remarks', '')
            
            # Update journal publication
            cursor.execute('''
                UPDATE journal_publications SET 
                department=%s, first_author=%s, corresponding_author=%s, other_authors=%s,
                faculty_author_position=%s, paper_title_apa=%s, journal_name=%s, volume_issue=%s,
                page_numbers=%s, issn_number=%s, doi=%s, year_of_publication=%s, indexing=%s,
                quartile=%s, impact_factor=%s, journal_link=%s, publisher=%s, funding_agency=%s, remarks=%s
                WHERE id=%s
            ''', (department, first_author, corresponding_author, other_authors,
                  faculty_author_position, paper_title_apa, journal_name, volume_issue,
                  page_numbers, issn_number, doi, year_of_publication, indexing, quartile,
                  impact_factor, journal_link, publisher, funding_agency, remarks, journal_id))
            
            conn.commit()
            cursor.close()
            conn.close()
            
            flash('✅ Journal publication updated successfully!', 'success')
            return redirect(f'/faculty/{journal["faculty_id"]}/publications')
        
        # GET request - show edit form
        cursor.close()
        conn.close()
        
        return render_template('edit_journal.html', journal=journal)
        
    except Exception as e:
        flash(f'❌ Error editing journal publication: {str(e)}', 'error')
        return redirect('/faculty')

# Edit Conference Publication
@app.route('/edit_conference/<int:conference_id>', methods=['GET', 'POST'])
@login_required
def edit_conference(conference_id):
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Get conference details
        cursor.execute('SELECT * FROM conference_publications WHERE id = %s', (conference_id,))
        conference = cursor.fetchone()
        
        if not conference:
            flash('❌ Conference publication not found!', 'error')
            return redirect('/faculty')
        
        # Check if user can edit this faculty's publications
        if not can_edit_publications(conference['faculty_id']):
            flash('❌ Access denied. You can only edit your own R&D publications.', 'error')
            return redirect(f'/faculty/{conference["faculty_id"]}/publications')
        
        if request.method == 'POST':
            # Get form data
            department = request.form['department']
            paper_title = request.form['paper_title']
            authors = request.form['authors']
            corresponding_author = request.form['corresponding_author']
            faculty_author_position = request.form['faculty_author_position']
            conference_name = request.form['conference_name']
            conference_venue = request.form.get('conference_venue', '')
            conference_dates = request.form.get('conference_dates', '')
            proceedings_title = request.form.get('proceedings_title', '')
            isbn_issn = request.form.get('isbn_issn', '')
            doi = request.form.get('doi', '')
            year_of_publication = int(request.form['year_of_publication'])
            indexing = request.form.get('indexing', '')
            publisher = request.form.get('publisher', '')
            conference_link = request.form.get('conference_link', '')
            
            # Update conference publication
            cursor.execute('''
                UPDATE conference_publications SET 
                department=%s, paper_title=%s, authors=%s, corresponding_author=%s,
                faculty_author_position=%s, conference_name=%s, conference_venue=%s, conference_dates=%s,
                proceedings_title=%s, isbn_issn=%s, doi=%s, year_of_publication=%s, indexing=%s,
                publisher=%s, conference_link=%s
                WHERE id=%s
            ''', (department, paper_title, authors, corresponding_author,
                  faculty_author_position, conference_name, conference_venue, conference_dates,
                  proceedings_title, isbn_issn, doi, year_of_publication, indexing,
                  publisher, conference_link, conference_id))
            
            conn.commit()
            cursor.close()
            conn.close()
            
            flash('✅ Conference publication updated successfully!', 'success')
            return redirect(f'/faculty/{conference["faculty_id"]}/publications')
        
        # GET request - show edit form
        cursor.close()
        conn.close()
        
        return render_template('edit_conference.html', conference=conference)
        
    except Exception as e:
        flash(f'❌ Error editing conference publication: {str(e)}', 'error')
        return redirect('/faculty')

# Edit Book Chapter
@app.route('/edit_book_chapter/<int:chapter_id>', methods=['GET', 'POST'])
@login_required
def edit_book_chapter(chapter_id):
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Get book chapter details
        cursor.execute('SELECT * FROM book_chapters WHERE id = %s', (chapter_id,))
        chapter = cursor.fetchone()
        
        if not chapter:
            flash('❌ Book chapter not found!', 'error')
            return redirect('/faculty')
        
        # Check if user can edit this faculty's publications
        if not can_edit_publications(chapter['faculty_id']):
            flash('❌ Access denied. You can only edit your own R&D publications.', 'error')
            return redirect(f'/faculty/{chapter["faculty_id"]}/publications')
        
        if request.method == 'POST':
            # Get form data
            department = request.form['department']
            chapter_title = request.form['chapter_title']
            book_title = request.form['book_title']
            authors = request.form['authors']
            faculty_author_position = request.form['faculty_author_position']
            corresponding_author = request.form['corresponding_author']
            publisher = request.form['publisher']
            isbn_number = request.form.get('isbn_number', '')
            chapter_doi = request.form.get('chapter_doi', '')
            year_of_publication = int(request.form['year_of_publication'])
            indexing = request.form.get('indexing', '')
            quartile = request.form.get('quartile', '')
            impact_factor = float(request.form.get('impact_factor', 0))
            chapter_link = request.form.get('chapter_link', '')
            
            # Update book chapter
            cursor.execute('''
                UPDATE book_chapters SET 
                department=%s, chapter_title=%s, book_title=%s, authors=%s,
                faculty_author_position=%s, corresponding_author=%s, publisher=%s, isbn_number=%s,
                chapter_doi=%s, year_of_publication=%s, indexing=%s, quartile=%s, impact_factor=%s, chapter_link=%s
                WHERE id=%s
            ''', (department, chapter_title, book_title, authors,
                  faculty_author_position, corresponding_author, publisher, isbn_number,
                  chapter_doi, year_of_publication, indexing, quartile, impact_factor, chapter_link, chapter_id))
            
            conn.commit()
            cursor.close()
            conn.close()
            
            flash('✅ Book chapter updated successfully!', 'success')
            return redirect(f'/faculty/{chapter["faculty_id"]}/publications')
        
        # GET request - show edit form
        cursor.close()
        conn.close()
        
        return render_template('edit_book_chapter.html', chapter=chapter)
        
    except Exception as e:
        flash(f'❌ Error editing book chapter: {str(e)}', 'error')
        return redirect('/faculty')

# Edit Patent
@app.route('/edit_patent/<int:patent_id>', methods=['GET', 'POST'])
@login_required
def edit_patent(patent_id):
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Get patent details
        cursor.execute('SELECT * FROM patents WHERE id = %s', (patent_id,))
        patent = cursor.fetchone()
        
        if not patent:
            flash('❌ Patent not found!', 'error')
            return redirect('/faculty')
        
        # Check if user can edit this faculty's publications
        if not can_edit_publications(patent['faculty_id']):
            flash('❌ Access denied. You can only edit your own R&D publications.', 'error')
            return redirect(f'/faculty/{patent["faculty_id"]}/publications')
        
        if request.method == 'POST':
            # Get form data
            department = request.form['department']
            patent_title = request.form['patent_title']
            inventors = request.form['inventors']
            corresponding_applicant = request.form['corresponding_applicant']
            faculty_author_position = request.form['faculty_author_position']
            patent_application_number = request.form['patent_application_number']
            filing_date = request.form.get('filing_date')
            publication_date = request.form.get('publication_date')
            grant_date = request.form.get('grant_date')
            patent_office = request.form['patent_office']
            status = request.form['status']
            patent_type = request.form['patent_type']
            patent_link = request.form.get('patent_link', '')
            certificate_link = request.form.get('certificate_link', '')
            
            # Convert empty dates to None
            filing_date = filing_date if filing_date else None
            publication_date = publication_date if publication_date else None
            grant_date = grant_date if grant_date else None
            
            # Update patent
            cursor.execute('''
                UPDATE patents SET 
                department=%s, patent_title=%s, inventors=%s, corresponding_applicant=%s,
                faculty_author_position=%s, patent_application_number=%s, filing_date=%s, publication_date=%s,
                grant_date=%s, patent_office=%s, status=%s, patent_type=%s, patent_link=%s, certificate_link=%s
                WHERE id=%s
            ''', (department, patent_title, inventors, corresponding_applicant,
                  faculty_author_position, patent_application_number, filing_date, publication_date,
                  grant_date, patent_office, status, patent_type, patent_link, certificate_link, patent_id))
            
            conn.commit()
            cursor.close()
            conn.close()
            
            flash('✅ Patent updated successfully!', 'success')
            return redirect(f'/faculty/{patent["faculty_id"]}/publications')
        
        # GET request - show edit form
        cursor.close()
        conn.close()
        
        return render_template('edit_patent.html', patent=patent)
        
    except Exception as e:
        flash(f'❌ Error editing patent: {str(e)}', 'error')
        return redirect('/faculty')

@app.route('/download_all_publications/<int:faculty_id>')
@login_required
def download_all_publications(faculty_id):
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Get faculty details
        cursor.execute('SELECT name_ssc, employee_id FROM faculty WHERE id = %s', (faculty_id,))
        faculty = cursor.fetchone()
        
        if not faculty:
            flash('❌ Faculty member not found!', 'error')
            return redirect('/faculty')
        
        # Get all publications data
        cursor.execute('SELECT * FROM journal_publications WHERE faculty_id = %s ORDER BY year_of_publication DESC', (faculty_id,))
        journals = cursor.fetchall()
        
        cursor.execute('SELECT * FROM conference_publications WHERE faculty_id = %s ORDER BY year_of_publication DESC', (faculty_id,))
        conferences = cursor.fetchall()
        
        cursor.execute('SELECT * FROM book_chapters WHERE faculty_id = %s ORDER BY year_of_publication DESC', (faculty_id,))
        book_chapters = cursor.fetchall()
        
        cursor.execute('SELECT * FROM patents WHERE faculty_id = %s ORDER BY filing_date DESC', (faculty_id,))
        patents = cursor.fetchall()
        
        cursor.close()
        conn.close()
        
        # Create Excel workbook with multiple sheets
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Add sheets for each publication type
        if journals:
            ws_journals = wb.create_sheet("Journal Publications")
            add_journals_to_sheet(ws_journals, journals)
        
        if conferences:
            ws_conferences = wb.create_sheet("Conference Papers")
            add_conferences_to_sheet(ws_conferences, conferences)
        
        if book_chapters:
            ws_book_chapters = wb.create_sheet("Book Chapters")
            add_book_chapters_to_sheet(ws_book_chapters, book_chapters)
        
        if patents:
            ws_patents = wb.create_sheet("Patents")
            add_patents_to_sheet(ws_patents, patents)
        
        # If no publications, create a message sheet
        if not any([journals, conferences, book_chapters, patents]):
            ws_empty = wb.create_sheet("No Publications")
            ws_empty['A1'] = "No R&D publications found for this faculty member."
        
        # Save to bytes buffer
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        # Create filename
        from datetime import datetime
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{faculty['employee_id']}_{faculty['name_ssc']}_All_Publications_{timestamp}.xlsx"
        
        return send_file(
            excel_buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        flash(f'❌ Error generating combined Excel file: {str(e)}', 'error')
        return redirect(f'/faculty/{faculty_id}/publications')

# Helper functions for each publication type
def add_journals_to_sheet(ws, journals):
    headers = ['S.No', 'Paper Title', 'Journal Name', 'First Author', 'Corresponding Author', 
               'Other Authors', 'Faculty Position', 'Year', 'Volume & Issue', 'Pages', 
               'ISSN', 'DOI', 'Indexing', 'Quartile', 'Impact Factor', 'Publisher']
    
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header).font = Font(bold=True)
    
    for row, journal in enumerate(journals, 2):
        ws.cell(row=row, column=1, value=row-1)
        ws.cell(row=row, column=2, value=journal['paper_title_apa'])
        ws.cell(row=row, column=3, value=journal['journal_name'])
        ws.cell(row=row, column=4, value=journal['first_author'])
        ws.cell(row=row, column=5, value=journal['corresponding_author'])
        ws.cell(row=row, column=6, value=journal['other_authors'] or '')
        ws.cell(row=row, column=7, value=journal['faculty_author_position'])
        ws.cell(row=row, column=8, value=journal['year_of_publication'])
        ws.cell(row=row, column=9, value=journal['volume_issue'] or '')
        ws.cell(row=row, column=10, value=journal['page_numbers'] or '')
        ws.cell(row=row, column=11, value=journal['issn_number'] or '')
        ws.cell(row=row, column=12, value=journal['doi'] or '')
        ws.cell(row=row, column=13, value=journal['indexing'] or '')
        ws.cell(row=row, column=14, value=journal['quartile'] or '')
        ws.cell(row=row, column=15, value=journal['impact_factor'] or '')
        ws.cell(row=row, column=16, value=journal['publisher'] or '')
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min((max_length + 2), 50)
        ws.column_dimensions[column_letter].width = adjusted_width

def add_conferences_to_sheet(ws, conferences):
    headers = ['S.No', 'Paper Title', 'Conference Name', 'Authors', 'Corresponding Author',
               'Faculty Position', 'Venue', 'Dates', 'Year', 'Proceedings', 'ISBN/ISSN', 'DOI']
    
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header).font = Font(bold=True)
    
    for row, conference in enumerate(conferences, 2):
        ws.cell(row=row, column=1, value=row-1)
        ws.cell(row=row, column=2, value=conference['paper_title'])
        ws.cell(row=row, column=3, value=conference['conference_name'])
        ws.cell(row=row, column=4, value=conference['authors'])
        ws.cell(row=row, column=5, value=conference['corresponding_author'])
        ws.cell(row=row, column=6, value=conference['faculty_author_position'])
        ws.cell(row=row, column=7, value=conference['conference_venue'] or '')
        ws.cell(row=row, column=8, value=conference['conference_dates'] or '')
        ws.cell(row=row, column=9, value=conference['year_of_publication'])
        ws.cell(row=row, column=10, value=conference['proceedings_title'] or '')
        ws.cell(row=row, column=11, value=conference['isbn_issn'] or '')
        ws.cell(row=row, column=12, value=conference['doi'] or '')
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min((max_length + 2), 50)
        ws.column_dimensions[column_letter].width = adjusted_width

def add_book_chapters_to_sheet(ws, chapters):
    headers = ['S.No', 'Chapter Title', 'Book Title', 'Authors', 'Corresponding Author',
               'Faculty Position', 'Publisher', 'ISBN', 'Year', 'DOI', 'Impact Factor']
    
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header).font = Font(bold=True)
    
    for row, chapter in enumerate(chapters, 2):
        ws.cell(row=row, column=1, value=row-1)
        ws.cell(row=row, column=2, value=chapter['chapter_title'])
        ws.cell(row=row, column=3, value=chapter['book_title'])
        ws.cell(row=row, column=4, value=chapter['authors'])
        ws.cell(row=row, column=5, value=chapter['corresponding_author'])
        ws.cell(row=row, column=6, value=chapter['faculty_author_position'])
        ws.cell(row=row, column=7, value=chapter['publisher'])
        ws.cell(row=row, column=8, value=chapter['isbn_number'] or '')
        ws.cell(row=row, column=9, value=chapter['year_of_publication'])
        ws.cell(row=row, column=10, value=chapter['chapter_doi'] or '')
        ws.cell(row=row, column=11, value=chapter['impact_factor'] or '')
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min((max_length + 2), 50)
        ws.column_dimensions[column_letter].width = adjusted_width

def add_patents_to_sheet(ws, patents):
    headers = ['S.No', 'Patent Title', 'Application Number', 'Inventors', 'Corresponding Applicant',
               'Faculty Position', 'Patent Office', 'Status', 'Type', 'Filing Date', 'Grant Date']
    
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header).font = Font(bold=True)
    
    for row, patent in enumerate(patents, 2):
        ws.cell(row=row, column=1, value=row-1)
        ws.cell(row=row, column=2, value=patent['patent_title'])
        ws.cell(row=row, column=3, value=patent['patent_application_number'])
        ws.cell(row=row, column=4, value=patent['inventors'])
        ws.cell(row=row, column=5, value=patent['corresponding_applicant'])
        ws.cell(row=row, column=6, value=patent['faculty_author_position'])
        ws.cell(row=row, column=7, value=patent['patent_office'])
        ws.cell(row=row, column=8, value=patent['status'])
        ws.cell(row=row, column=9, value=patent['patent_type'])
        ws.cell(row=row, column=10, value=str(patent['filing_date']) if patent['filing_date'] else '')
        ws.cell(row=row, column=11, value=str(patent['grant_date']) if patent['grant_date'] else '')
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min((max_length + 2), 50)
        ws.column_dimensions[column_letter].width = adjusted_width        

@app.route('/edit_qualification/<int:qualification_id>', methods=['GET', 'POST'])
@login_required
def edit_qualification(qualification_id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    
    # Get qualification details
    cursor.execute('SELECT * FROM qualifications WHERE id = %s', (qualification_id,))
    qualification = cursor.fetchone()
    
    if not qualification:
        flash('❌ Qualification not found!', 'error')
        return redirect('/faculty')
    
    # Get faculty details for navigation
    cursor.execute('SELECT * FROM faculty WHERE id = %s', (qualification['faculty_id'],))
    faculty = cursor.fetchone()
    
    if request.method == 'POST':
        # Update qualification
        qualification_type = request.form['qualification_type']
        domain_specialization = request.form.get('domain_specialization', '')
        percentage = request.form.get('percentage', '')
        year_of_passing = request.form['year_of_passing']
        institution_name = request.form['institution_name']
        highest_degree = 'highest_degree' in request.form
        pursuing = 'pursuing' in request.form
        
        cursor.execute('''
            UPDATE qualifications SET 
            qualification_type=%s, domain_specialization=%s, percentage=%s, 
            year_of_passing=%s, institution_name=%s, highest_degree=%s, pursuing=%s
            WHERE id=%s
        ''', (qualification_type, domain_specialization, percentage, 
              year_of_passing, institution_name, highest_degree, pursuing, qualification_id))
        
        conn.commit()
        cursor.close()
        conn.close()
        
        flash('✅ Qualification updated successfully!', 'success')
        return redirect(f'/faculty/{qualification["faculty_id"]}/qualifications')
    
    cursor.close()
    conn.close()
    return render_template('edit_qualification.html', qualification=qualification, faculty=faculty)

@app.route('/download_qualifications/<int:faculty_id>')
@login_required
def download_qualifications(faculty_id):
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Get faculty details
        cursor.execute('SELECT * FROM faculty WHERE id = %s', (faculty_id,))
        faculty = cursor.fetchone()
        
        # Get all qualifications
        cursor.execute('SELECT * FROM qualifications WHERE faculty_id = %s ORDER BY year_of_passing DESC', (faculty_id,))
        qualifications = cursor.fetchall()
        
        cursor.close()
        conn.close()
        
        if not faculty:
            flash('❌ Faculty member not found!', 'error')
            return redirect('/faculty')
        
        # Create Excel file
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Qualifications"
        
        # Headers
        headers = ['S.No', 'Qualification', 'Specialization', 'Percentage', 'Year', 'Institution', 'Status']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header).font = Font(bold=True)
        
        # Data
        for row, qual in enumerate(qualifications, 2):
            ws.cell(row=row, column=1, value=row-1)
            ws.cell(row=row, column=2, value=qual['qualification_type'])
            ws.cell(row=row, column=3, value=qual['domain_specialization'] or '')
            ws.cell(row=row, column=4, value=qual['percentage'] or '')
            ws.cell(row=row, column=5, value=qual['year_of_passing'])
            ws.cell(row=row, column=6, value=qual['institution_name'])
            status = []
            if qual['highest_degree']: status.append('Highest Degree')
            if qual['pursuing']: status.append('Pursuing')
            ws.cell(row=row, column=7, value=', '.join(status) if status else 'Completed')
        
        # Auto-adjust columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except: pass
            adjusted_width = min((max_length + 2), 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to bytes buffer
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        filename = f"qualifications_{faculty['employee_id']}_{faculty['name_ssc']}.xlsx"
        
        return send_file(
            excel_buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        flash(f'❌ Error downloading qualifications: {str(e)}', 'error')
        return redirect(f'/faculty/{faculty_id}/qualifications')
@app.route('/check_faculty_access')
@login_required
def check_faculty_access():
    """Simple access check for faculty users for department/experience"""
    try:
        access_type = request.args.get('type', '')
        user_role = get_user_role()
        user_email = session.get('email', '')
        
        print(f"🔍 ACCESS CHECK: role='{user_role}', type='{access_type}', email='{user_email}'")
        
        # IQAC and Office always go to faculty list
        if user_role in ['IQAC', 'Office']:
            if access_type == 'department':
                return jsonify({
                    'access_granted': True,
                    'redirect_url': '/faculty?view=departments',
                    'message': '✅ Access granted - Viewing all departments'
                })
            elif access_type == 'experience':
                return jsonify({
                    'access_granted': True,
                    'redirect_url': '/faculty?view=experience',
                    'message': '✅ Access granted - Viewing all experience levels'
                })
            else:
                return jsonify({
                    'access_granted': True,
                    'redirect_url': '/faculty',
                    'message': '✅ Access granted - Viewing all faculty'
                })
        
        # For Faculty users, find their profile and redirect appropriately
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute('SELECT * FROM faculty WHERE email = %s', (user_email,))
        faculty_profile = cursor.fetchone()
        cursor.close()
        conn.close()
        
        if not faculty_profile:
            return jsonify({
                'access_granted': False,
                'message': '❌ Access Denied: You do not have a faculty profile. Please contact administrator.',
                'redirect_url': '/faculty'
            })
        
        # Determine redirect URL based on access type
        if access_type == 'department':
            redirect_url = f'/department/{faculty_profile["department"]}'
            message = f'✅ Access granted - Viewing your department: {faculty_profile["department"]}'
        elif access_type == 'experience':
            experience_cat = faculty_profile.get('experience_category', '0-5')
            redirect_url = f'/experience/{experience_cat}'
            message = f'✅ Access granted - Viewing your experience category: {experience_cat} years'
        else:
            redirect_url = '/faculty'
            message = '✅ Access granted - Viewing your profile'
        
        return jsonify({
            'access_granted': True,
            'redirect_url': redirect_url,
            'message': message
        })
        
    except Exception as e:
        print(f"❌ ERROR in check_faculty_access: {str(e)}")
        return jsonify({
            'access_granted': False,
            'message': f'❌ System error: {str(e)}',
            'redirect_url': '/'
        })
@app.route('/check_designation_access')
@login_required
def check_designation_access():
    """Check if user has access to view a specific designation"""
    try:
        requested_designation = request.args.get('designation', '')
        user_role = get_user_role()
        user_email = session.get('email', '')
        
        print(f"🔍 DESIGNATION ACCESS CHECK: role='{user_role}', designation='{requested_designation}', email='{user_email}'")
        
        # IQAC and Office can view any designation
        if user_role in ['IQAC', 'Office']:
            return jsonify({
                'access_granted': True,
                'redirect_url': f'/designation/{requested_designation}',
                'message': f'✅ Access granted - Viewing {requested_designation} designation'
            })
        
        # For Faculty users, check if this is their designation
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute('SELECT designation FROM faculty WHERE email = %s', (user_email,))
        faculty_profile = cursor.fetchone()
        cursor.close()
        conn.close()
        
        if not faculty_profile:
            return jsonify({
                'access_granted': False,
                'message': '❌ Access Denied: You do not have a faculty profile.',
                'redirect_url': '/faculty'
            })
        
        user_designation = faculty_profile['designation']
        
        # Check if user is trying to access their own designation
        if user_designation == requested_designation:
            return jsonify({
                'access_granted': True,
                'redirect_url': f'/designation/{requested_designation}',
                'message': f'✅ Access granted - Viewing your designation: {requested_designation}'
            })
        else:
            return jsonify({
                'access_granted': False,
                'message': f'❌ Access Denied: You can only view your own designation ({user_designation}), not {requested_designation}.',
                'redirect_url': f'/designation/{user_designation}'
            })
        
    except Exception as e:
        print(f"❌ ERROR in check_designation_access: {str(e)}")
        return jsonify({
            'access_granted': False,
            'message': f'❌ System error: {str(e)}',
            'redirect_url': '/faculty'
        })    
# =====================
# R&D PUBLICATIONS MASTER VIEW ROUTES
# =====================
@app.route('/rd/publications')
@login_required
def rd_publications_master():
    """Master view for all R&D publications with filters - for IQAC/Office only"""
    if get_user_role() not in ['IQAC', 'Office']:
        flash('❌ Access denied. IQAC/Office privileges required.', 'error')
        return redirect('/')
    
    publication_type = request.args.get('type', 'journal')
    department = request.args.get('department', '')
    year = request.args.get('year', '')
    indexing = request.args.get('indexing', '')
    status = request.args.get('status', '')
    
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    
    publications = []
    stats = {}
    
    # Define departments list
    departments = ['CIVIL', 'EEE', 'MECH', 'ECE', 'CSE', 'CSE-AI', 'CSE-DS', 'CSE-AI&ML', 'CSE-CS', 'IT']
    
    # Define indexing options
    indexing_options = ['Scopus', 'SCI', 'SCIE', 'WoS']
    
    # Define patent status options
    patent_statuses = ['Filed', 'Published', 'Granted']
    
    if publication_type == 'journal':
        # Build query for journals
        query = '''
            SELECT j.*, f.name_ssc, f.department as faculty_department, f.employee_id 
            FROM journal_publications j 
            JOIN faculty f ON j.faculty_id = f.id 
            WHERE 1=1
        '''
        params = []
        
        if department:
            query += ' AND j.department = %s'
            params.append(department)
        if year:
            query += ' AND j.year_of_publication = %s'
            params.append(year)
        if indexing:
            query += ' AND j.indexing = %s'
            params.append(indexing)
        
        query += ' ORDER BY j.year_of_publication DESC, j.department'
        
        cursor.execute(query, params)
        publications = cursor.fetchall()
        
        # Get stats
        cursor.execute('SELECT COUNT(*) as total FROM journal_publications')
        stats['total'] = cursor.fetchone()['total']
        
    elif publication_type == 'conference':
        # Build query for conferences
        query = '''
            SELECT c.*, f.name_ssc, f.department as faculty_department, f.employee_id 
            FROM conference_publications c 
            JOIN faculty f ON c.faculty_id = f.id 
            WHERE 1=1
        '''
        params = []
        
        if department:
            query += ' AND c.department = %s'
            params.append(department)
        if year:
            query += ' AND c.year_of_publication = %s'
            params.append(year)
        
        query += ' ORDER BY c.year_of_publication DESC, c.department'
        
        cursor.execute(query, params)
        publications = cursor.fetchall()
        
        # Get stats
        cursor.execute('SELECT COUNT(*) as total FROM conference_publications')
        stats['total'] = cursor.fetchone()['total']
        
    elif publication_type == 'book_chapter':
        # Build query for book chapters
        query = '''
            SELECT b.*, f.name_ssc, f.department as faculty_department, f.employee_id 
            FROM book_chapters b 
            JOIN faculty f ON b.faculty_id = f.id 
            WHERE 1=1
        '''
        params = []
        
        if department:
            query += ' AND b.department = %s'
            params.append(department)
        if year:
            query += ' AND b.year_of_publication = %s'
            params.append(year)
        
        query += ' ORDER BY b.year_of_publication DESC, b.department'
        
        cursor.execute(query, params)
        publications = cursor.fetchall()
        
        # Get stats
        cursor.execute('SELECT COUNT(*) as total FROM book_chapters')
        stats['total'] = cursor.fetchone()['total']
        
    elif publication_type == 'patent':
        # Build query for patents
        query = '''
            SELECT p.*, f.name_ssc, f.department as faculty_department, f.employee_id 
            FROM patents p 
            JOIN faculty f ON p.faculty_id = f.id 
            WHERE 1=1
        '''
        params = []
        
        if department:
            query += ' AND p.department = %s'
            params.append(department)
        if year:
            query += ' AND YEAR(p.filing_date) = %s'
            params.append(year)
        if status:
            query += ' AND p.status = %s'
            params.append(status)
        
        query += ' ORDER BY p.filing_date DESC, p.department'
        
        cursor.execute(query, params)
        publications = cursor.fetchall()
        
        # Get stats
        cursor.execute('SELECT COUNT(*) as total FROM patents')
        stats['total'] = cursor.fetchone()['total']
    
    cursor.close()
    conn.close()
    
    # Get current year and last 10 years for year filter
    current_year = datetime.datetime.now().year
    years = list(range(current_year, current_year - 10, -1))
    
    return render_template('rd_publications_master.html',
                         publication_type=publication_type,
                         publications=publications,
                         departments=departments,
                         years=years,
                         indexings=indexing_options,
                         patent_statuses=patent_statuses,
                         selected_department=department,
                         selected_year=year,
                         selected_indexing=indexing,
                         selected_status=status,
                         stats=stats)

@app.route('/rd/download_excel')
@login_required
def rd_download_excel():
    """Download R&D publications as Excel - for IQAC/Office only"""
    if get_user_role() not in ['IQAC', 'Office']:
        flash('❌ Access denied. IQAC/Office privileges required.', 'error')
        return redirect('/')
    
    try:
        publication_type = request.args.get('type', 'journal')
        department = request.args.get('department', '')
        year = request.args.get('year', '')
        indexing = request.args.get('indexing', '')
        status = request.args.get('status', '')
        
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        if publication_type == 'journal':
            query = '''
                SELECT j.*, f.name_ssc, f.department as faculty_department, f.employee_id 
                FROM journal_publications j 
                JOIN faculty f ON j.faculty_id = f.id 
                WHERE 1=1
            '''
            params = []
            
            if department:
                query += ' AND j.department = %s'
                params.append(department)
            if year:
                query += ' AND j.year_of_publication = %s'
                params.append(year)
            if indexing:
                query += ' AND j.indexing = %s'
                params.append(indexing)
            
            query += ' ORDER BY j.year_of_publication DESC, j.department'
            
            cursor.execute(query, params)
            publications = cursor.fetchall()
            
            # Create Excel for journals
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Journal Publications"
            
            headers = ['S.No', 'Employee ID', 'Faculty Name', 'Department', 'Paper Title', 
                      'Journal Name', 'First Author', 'Corresponding Author', 'Year', 
                      'Volume & Issue', 'Pages', 'ISSN', 'DOI', 'Indexing', 'Quartile', 
                      'Impact Factor', 'Publisher']
            
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header).font = Font(bold=True)
            
            for row, pub in enumerate(publications, 2):
                ws.cell(row=row, column=1, value=row-1)
                ws.cell(row=row, column=2, value=pub['employee_id'])
                ws.cell(row=row, column=3, value=pub['name_ssc'])
                ws.cell(row=row, column=4, value=pub['department'])
                ws.cell(row=row, column=5, value=pub['paper_title_apa'])
                ws.cell(row=row, column=6, value=pub['journal_name'])
                ws.cell(row=row, column=7, value=pub['first_author'])
                ws.cell(row=row, column=8, value=pub['corresponding_author'])
                ws.cell(row=row, column=9, value=pub['year_of_publication'])
                ws.cell(row=row, column=10, value=pub['volume_issue'] or '')
                ws.cell(row=row, column=11, value=pub['page_numbers'] or '')
                ws.cell(row=row, column=12, value=pub['issn_number'] or '')
                ws.cell(row=row, column=13, value=pub['doi'] or '')
                ws.cell(row=row, column=14, value=pub['indexing'] or '')
                ws.cell(row=row, column=15, value=pub['quartile'] or '')
                ws.cell(row=row, column=16, value=pub['impact_factor'] or '')
                ws.cell(row=row, column=17, value=pub['publisher'] or '')
                
        elif publication_type == 'conference':
            query = '''
                SELECT c.*, f.name_ssc, f.department as faculty_department, f.employee_id 
                FROM conference_publications c 
                JOIN faculty f ON c.faculty_id = f.id 
                WHERE 1=1
            '''
            params = []
            
            if department:
                query += ' AND c.department = %s'
                params.append(department)
            if year:
                query += ' AND c.year_of_publication = %s'
                params.append(year)
            
            query += ' ORDER BY c.year_of_publication DESC, c.department'
            
            cursor.execute(query, params)
            publications = cursor.fetchall()
            
            # Create Excel for conferences
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Conference Publications"
            
            headers = ['S.No', 'Employee ID', 'Faculty Name', 'Department', 'Paper Title', 
                      'Conference Name', 'Authors', 'Corresponding Author', 'Year', 
                      'Venue', 'Dates', 'Proceedings', 'ISBN/ISSN', 'DOI', 'Indexing']
            
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header).font = Font(bold=True)
            
            for row, pub in enumerate(publications, 2):
                ws.cell(row=row, column=1, value=row-1)
                ws.cell(row=row, column=2, value=pub['employee_id'])
                ws.cell(row=row, column=3, value=pub['name_ssc'])
                ws.cell(row=row, column=4, value=pub['department'])
                ws.cell(row=row, column=5, value=pub['paper_title'])
                ws.cell(row=row, column=6, value=pub['conference_name'])
                ws.cell(row=row, column=7, value=pub['authors'])
                ws.cell(row=row, column=8, value=pub['corresponding_author'])
                ws.cell(row=row, column=9, value=pub['year_of_publication'])
                ws.cell(row=row, column=10, value=pub['conference_venue'] or '')
                ws.cell(row=row, column=11, value=pub['conference_dates'] or '')
                ws.cell(row=row, column=12, value=pub['proceedings_title'] or '')
                ws.cell(row=row, column=13, value=pub['isbn_issn'] or '')
                ws.cell(row=row, column=14, value=pub['doi'] or '')
                ws.cell(row=row, column=15, value=pub['indexing'] or '')
                
        elif publication_type == 'book_chapter':
            query = '''
                SELECT b.*, f.name_ssc, f.department as faculty_department, f.employee_id 
                FROM book_chapters b 
                JOIN faculty f ON b.faculty_id = f.id 
                WHERE 1=1
            '''
            params = []
            
            if department:
                query += ' AND b.department = %s'
                params.append(department)
            if year:
                query += ' AND b.year_of_publication = %s'
                params.append(year)
            
            query += ' ORDER BY b.year_of_publication DESC, b.department'
            
            cursor.execute(query, params)
            publications = cursor.fetchall()
            
            # Create Excel for book chapters
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Book Chapters"
            
            headers = ['S.No', 'Employee ID', 'Faculty Name', 'Department', 'Chapter Title', 
                      'Book Title', 'Authors', 'Corresponding Author', 'Publisher', 
                      'ISBN', 'Year', 'Chapter DOI', 'Indexing', 'Impact Factor']
            
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header).font = Font(bold=True)
            
            for row, pub in enumerate(publications, 2):
                ws.cell(row=row, column=1, value=row-1)
                ws.cell(row=row, column=2, value=pub['employee_id'])
                ws.cell(row=row, column=3, value=pub['name_ssc'])
                ws.cell(row=row, column=4, value=pub['department'])
                ws.cell(row=row, column=5, value=pub['chapter_title'])
                ws.cell(row=row, column=6, value=pub['book_title'])
                ws.cell(row=row, column=7, value=pub['authors'])
                ws.cell(row=row, column=8, value=pub['corresponding_author'])
                ws.cell(row=row, column=9, value=pub['publisher'])
                ws.cell(row=row, column=10, value=pub['isbn_number'] or '')
                ws.cell(row=row, column=11, value=pub['year_of_publication'])
                ws.cell(row=row, column=12, value=pub['chapter_doi'] or '')
                ws.cell(row=row, column=13, value=pub['indexing'] or '')
                ws.cell(row=row, column=14, value=pub['impact_factor'] or '')
                
        elif publication_type == 'patent':
            query = '''
                SELECT p.*, f.name_ssc, f.department as faculty_department, f.employee_id 
                FROM patents p 
                JOIN faculty f ON p.faculty_id = f.id 
                WHERE 1=1
            '''
            params = []
            
            if department:
                query += ' AND p.department = %s'
                params.append(department)
            if year:
                query += ' AND YEAR(p.filing_date) = %s'
                params.append(year)
            if status:
                query += ' AND p.status = %s'
                params.append(status)
            
            query += ' ORDER BY p.filing_date DESC, p.department'
            
            cursor.execute(query, params)
            publications = cursor.fetchall()
            
            # Create Excel for patents
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Patents"
            
            headers = ['S.No', 'Employee ID', 'Faculty Name', 'Department', 'Patent Title', 
                      'Application Number', 'Inventors', 'Corresponding Applicant', 
                      'Patent Office', 'Status', 'Type', 'Filing Date', 'Publication Date', 
                      'Grant Date']
            
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header).font = Font(bold=True)
            
            for row, pub in enumerate(publications, 2):
                ws.cell(row=row, column=1, value=row-1)
                ws.cell(row=row, column=2, value=pub['employee_id'])
                ws.cell(row=row, column=3, value=pub['name_ssc'])
                ws.cell(row=row, column=4, value=pub['department'])
                ws.cell(row=row, column=5, value=pub['patent_title'])
                ws.cell(row=row, column=6, value=pub['patent_application_number'])
                ws.cell(row=row, column=7, value=pub['inventors'])
                ws.cell(row=row, column=8, value=pub['corresponding_applicant'])
                ws.cell(row=row, column=9, value=pub['patent_office'])
                ws.cell(row=row, column=10, value=pub['status'])
                ws.cell(row=row, column=11, value=pub['patent_type'])
                ws.cell(row=row, column=12, value=str(pub['filing_date']) if pub['filing_date'] else '')
                ws.cell(row=row, column=13, value=str(pub['publication_date']) if pub['publication_date'] else '')
                ws.cell(row=row, column=14, value=str(pub['grant_date']) if pub['grant_date'] else '')
        
        cursor.close()
        conn.close()
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min((max_length + 2), 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to bytes buffer
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        # Create filename
        from datetime import datetime
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"rd_{publication_type}_publications_{timestamp}.xlsx"
        
        return send_file(
            excel_buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        flash(f'❌ Error generating Excel file: {str(e)}', 'error')
        return redirect(f'/rd/publications?type={publication_type}')   
@app.route('/debug-db')
def debug_db():
    """Debug database connection"""
    try:
        env_info = {
            'MYSQLHOST': os.environ.get('MYSQLHOST'),
            'MYSQLUSER': os.environ.get('MYSQLUSER'),
            'MYSQLDATABASE': os.environ.get('MYSQLDATABASE'),
            'MYSQLPORT': os.environ.get('MYSQLPORT'),
            'MYSQL_URL': os.environ.get('MYSQL_URL')[:50] + '...' if os.environ.get('MYSQL_URL') else None
        }
        
        conn = get_db_connection()
        if conn:
            cursor = conn.cursor()
            cursor.execute("SELECT 1 as test")
            result = cursor.fetchone()
            cursor.close()
            conn.close()
            return jsonify({
                "status": "success", 
                "database_test": result,
                "environment": env_info
            })
        else:
            return jsonify({
                "status": "error", 
                "message": "Database connection failed",
                "environment": env_info
            })
    except Exception as e:
        return jsonify({
            "status": "error", 
            "message": str(e),
            "environment": env_info
        })                 
if __name__ == '__main__':
    print("🚀 Faculty Portal Starting...")
    app.run(debug=True, host='0.0.0.0', port=5000)